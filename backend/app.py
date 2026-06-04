import os
import csv
import io
import uuid
import traceback
from functools import wraps
from importlib import import_module
from flask import Flask, request, jsonify, send_from_directory, make_response, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from werkzeug.utils import secure_filename
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity, get_jwt
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

from datetime import timedelta, date, datetime
from zoneinfo import ZoneInfo
import json
from dotenv import load_dotenv

load_dotenv()

from database.database import get_db_connection, close_connection
from database.sync_pipeline import fill_pipeline
from gsheets_sync import sync_from_sheets

def ensure_schema():
    conn = get_db_connection()
    if not conn:
        print("Schema verification failed: could not connect to database.")
        return
    try:
        cursor = conn.cursor()
        # Ensure lost_reason column exists in deals table
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'deals' AND COLUMN_NAME = 'lost_reason'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'lost_reason' column to 'deals' table...")
            cursor.execute("ALTER TABLE deals ADD COLUMN lost_reason VARCHAR(255) NULL")
            conn.commit()

        # Ensure probability_manual column exists in deals table
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'deals' AND COLUMN_NAME = 'probability_manual'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'probability_manual' column to 'deals' table...")
            cursor.execute("ALTER TABLE deals ADD COLUMN probability_manual BOOLEAN DEFAULT FALSE")
            conn.commit()

        # Ensure user_id column exists in audit_log table
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'audit_log' AND COLUMN_NAME = 'user_id'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'user_id' column to 'audit_log' table...")
            cursor.execute("ALTER TABLE audit_log ADD COLUMN user_id INT NULL")
            cursor.execute("ALTER TABLE audit_log ADD FOREIGN KEY (user_id) REFERENCES team(id) ON DELETE SET NULL")
            conn.commit()

        # Ensure metadata column exists in activities table
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'activities' AND COLUMN_NAME = 'metadata'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'metadata' column to 'activities' table...")
            cursor.execute("ALTER TABLE activities ADD COLUMN metadata TEXT NULL")
            conn.commit()

        # Ensure deal_contacts table exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS deal_contacts (
                deal_id    VARCHAR(100) NOT NULL,
                contact_id VARCHAR(100) NOT NULL,
                role       VARCHAR(100) DEFAULT 'Primary',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                PRIMARY KEY (deal_id, contact_id),
                FOREIGN KEY (deal_id) REFERENCES deals(id) ON DELETE CASCADE,
                FOREIGN KEY (contact_id) REFERENCES contacts(id) ON DELETE CASCADE
            )
        """)
        conn.commit()

        # Ensure owner_name column exists in leads table
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'leads' AND COLUMN_NAME = 'owner_name'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'owner_name' column to 'leads' table...")
            cursor.execute("ALTER TABLE leads ADD COLUMN owner_name VARCHAR(255) NULL")
            conn.commit()

        # Ensure reassigned_at column exists in leads table (newly-handed-over "New" flag)
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'leads' AND COLUMN_NAME = 'reassigned_at'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'reassigned_at' column to 'leads' table...")
            cursor.execute("ALTER TABLE leads ADD COLUMN reassigned_at DATE NULL")
            conn.commit()

        # Ensure celebration_music table exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS celebration_music (
                id INT AUTO_INCREMENT PRIMARY KEY,
                outcome ENUM('won','lost') NOT NULL,
                source_type ENUM('url','internal') NOT NULL DEFAULT 'url',
                url VARCHAR(500) NOT NULL,
                original_filename VARCHAR(255),
                stored_filename VARCHAR(255),
                is_active TINYINT(1) DEFAULT 1,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

        # Ensure app_settings key-value table exists
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS app_settings (
                setting_key   VARCHAR(100) PRIMARY KEY,
                setting_value TEXT,
                updated_at    TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        """)
        conn.commit()

        # Ensure profile_pic column exists in team table
        cursor.execute("""
            SELECT COUNT(*) FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'team' AND COLUMN_NAME = 'profile_pic'
        """)
        if cursor.fetchone()[0] == 0:
            print("Adding missing 'profile_pic' column to 'team' table...")
            cursor.execute("ALTER TABLE team ADD COLUMN profile_pic VARCHAR(500) NULL")
            conn.commit()

        # Ensure adjustment columns exist
        cursor.execute("""
            SELECT COLUMN_NAME FROM information_schema.COLUMNS
            WHERE TABLE_SCHEMA = database() AND TABLE_NAME = 'team'
        """)
        existing_columns = [row[0] for row in cursor.fetchall()]

        if 'profile_zoom' not in existing_columns:
            print("Adding missing 'profile_zoom' column to 'team' table...")
            cursor.execute("ALTER TABLE team ADD COLUMN profile_zoom FLOAT DEFAULT 1.0")
        
        if 'profile_offset_y' not in existing_columns:
            print("Adding missing 'profile_offset_y' column to 'team' table...")
            cursor.execute("ALTER TABLE team ADD COLUMN profile_offset_y FLOAT DEFAULT 0.0")

        if 'profile_offset_x' not in existing_columns:
            print("Adding missing 'profile_offset_x' column to 'team' table...")
            cursor.execute("ALTER TABLE team ADD COLUMN profile_offset_x FLOAT DEFAULT 0.0")

        if 'profile_rotation' not in existing_columns:
            print("Adding missing 'profile_rotation' column to 'team' table...")
            cursor.execute("ALTER TABLE team ADD COLUMN profile_rotation INT DEFAULT 0")
        
        conn.commit()

        # Migrate 'Sales Rep' role → 'Branch Account'
        cursor.execute("UPDATE team SET role = 'Branch Account' WHERE role = 'Sales Rep'")
        conn.commit()
    except Exception as e:
        print(f"Error during schema verification: {e}")
    finally:
        close_connection(conn)

ensure_schema()

app = Flask(__name__)
CORS(app, resources={r"/api/*": {
    "origins": [os.getenv("FRONTEND_URL", "http://localhost:5173")],
    "allow_headers": ["Content-Type", "Authorization"]
}})

app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'super-secret-key')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=8)  # 8-hour sessions
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB upload limit
jwt = JWTManager(app)

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

MUSIC_UPLOAD_FOLDER = os.path.join(UPLOAD_FOLDER, 'celebration-music')
os.makedirs(MUSIC_UPLOAD_FOLDER, exist_ok=True)

PROFILE_PIC_FOLDER = os.path.join(UPLOAD_FOLDER, 'profile-pics')
os.makedirs(PROFILE_PIC_FOLDER, exist_ok=True)

ALLOWED_AUDIO_EXTENSIONS = {'.mp3', '.wav', '.ogg', '.m4a', '.aac', '.webm'}
ALLOWED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.gif', '.webp'}

REGION_BRANCHES = {
    'Central':     ['Manila', 'Palawan', 'Legazpi', 'Cavite', 'Batangas'],
    'North Luzon': ['Ilocos', 'Isabela'],
    'Vis&Min':     ['Gensan', 'Iloilo', 'Cebu', 'Davao', 'CDO'],
}

@jwt.unauthorized_loader
def unauthorized_response(callback):
    return jsonify({'error': 'Missing Authorization Header', 'details': callback}), 401

@jwt.invalid_token_loader
def invalid_token_response(callback):
    return jsonify({'error': 'Invalid Token', 'details': callback}), 401

@jwt.expired_token_loader
def expired_token_response(jwt_header, jwt_payload):
    return jsonify({'error': 'Token Expired'}), 401

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["1000 per day", "200 per hour"]
)

# Admin-only decorator: use on top of @jwt_required()
# Usage: @jwt_required() THEN @admin_required on the route
def admin_required(fn):
    """Decorator that checks the JWT claims for Admin role.
    Must be stacked BELOW @jwt_required() so the token is already validated.
    """
    @wraps(fn)
    def wrapper(*args, **kwargs):
        claims = get_jwt()
        if claims.get('role') != 'Admin':
            return jsonify(error='Admins only!'), 403
        return fn(*args, **kwargs)
    return wrapper

@app.errorhandler(Exception)
def handle_exception(e):
    from werkzeug.exceptions import HTTPException
    if isinstance(e, HTTPException):
        return e
    # Log securely on backend
    print(f"Backend Error: {traceback.format_exc()}")
    return jsonify(error="An internal error occurred"), 500

STAGE_PROBABILITY = {
    'Qualified':       20,
    'New Opportunity': 40,
    'Proposal':        60,
    'Negotiation':     80,
    'Closed Won':      100,
    'Closed Lost':     0,
}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def rows_to_list(cursor):
    """Return all rows from the cursor as a list of dicts."""
    columns = [col[0] for col in cursor.description]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def verify_password(stored_hash, provided_password):
    """Verify passwords across both Werkzeug and bcrypt hash formats.

    Returns:
        tuple[bool, bool]: (is_valid, should_rehash)
    """
    if not stored_hash or not provided_password:
        return False, False

    hash_value = stored_hash.decode('utf-8', errors='ignore') if isinstance(stored_hash, bytes) else str(stored_hash)
    password_value = str(provided_password)

    try:
        return check_password_hash(hash_value, password_value), False
    except (ValueError, TypeError):
        pass

    if hash_value.startswith(('$2a$', '$2b$', '$2y$')):
        try:
            bcrypt_module = import_module('bcrypt')
            is_valid = bcrypt_module.checkpw(password_value.encode('utf-8'), hash_value.encode('utf-8'))
            return is_valid, is_valid
        except (ImportError, ValueError, TypeError):
            return False, False

    return False, False


def normalize_username(username):
    return str(username or '').strip().lower()


def normalize_branch(branch):
    return str(branch or '').strip().lower()


def _parse_contact_name(raw):
    """Return a non-empty string if raw is a real name; return '' for None, 0, 0.0, '0.00', etc."""
    if raw is None:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    try:
        if float(s.replace(',', '')) == 0:
            return ''
    except ValueError:
        pass
    return s


def _parse_phone(raw):
    """Clean a phone value from Excel: strip whitespace, convert numeric floats to int strings."""
    if raw is None:
        return ''
    s = str(raw).strip()
    # Excel stores phone numbers as floats (e.g. 93120000.0) — strip the trailing .0
    if s.endswith('.0'):
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    return s if s not in ('0', '0.0', '0.00', '') else ''


def has_branch_filter(branch):
    normalized = normalize_branch(branch)
    return bool(normalized and normalized != 'headquarters')


ROLE_RANK = {
    'Admin': 100,
    'Head of Sales': 80,
    'Regional Sales Manager': 60,
    'Sales Representative': 40,
    'Branch Account': 40,
    'Sales Rep': 40,
}


def can_assign(claims, current_user_id, target_id, target_role):
    """
    Check if current user can assign to target_role (strictly below).
    Self-assignment is always allowed.
    """
    # Self-assignment is always allowed
    if str(target_id) == str(current_user_id):
        return True
    
    user_role = claims.get('role', '')
    user_rank = ROLE_RANK.get(user_role, 0)
    target_rank = ROLE_RANK.get(target_role, 0)
    
    # Admin can assign to anyone
    if user_role == 'Admin':
        return True
    
    return user_rank > target_rank


def build_scope(claims, requested_branch, col='LOWER(TRIM(l.branch))', requested_region=''):
    """
    Returns (where_parts, restrict_owner, params) for role-based data filtering.
    where_parts: SQL fragment list using `col` for branch matching.
    restrict_owner: True for Sales Rep/Representative — callers add own-alias owner_id filter.
    params: positional values for where_parts.
    """
    role        = claims.get('role', 'Branch Account')
    user_region = claims.get('region', '')
    user_branch = claims.get('branch', '')

    if role == 'Sales Representative':
        return ([f'{col} = %s'], True, [normalize_branch(user_branch)])

    if role in ('Branch Account', 'Sales Rep'):
        return ([f'{col} = %s'], False, [normalize_branch(user_branch)])

    if role == 'Regional Sales Manager':
        region_branches = REGION_BRANCHES.get(user_region, [])
        allowed = [normalize_branch(b) for b in region_branches]
        req = normalize_branch(requested_branch)
        if req and req in allowed:
            return ([f'{col} = %s'], False, [req])
        if allowed:
            ph = ', '.join(['%s'] * len(allowed))
            return ([f'{col} IN ({ph})'], False, allowed)
        return (['1=0'], False, [])

    if role in ('Head of Sales', 'Admin'):
        if has_branch_filter(requested_branch):
            return ([f'{col} = %s'], False, [normalize_branch(requested_branch)])
        if requested_region and requested_region in REGION_BRANCHES:
            region_branches = [normalize_branch(b) for b in REGION_BRANCHES[requested_region]]
            ph = ', '.join(['%s'] * len(region_branches))
            return ([f'{col} IN ({ph})'], False, region_branches)
        return ([], False, [])

    # Default: branch accounts ('Branch Account') and any unrecognised role
    if has_branch_filter(requested_branch):
        return ([f'{col} = %s'], False, [normalize_branch(requested_branch)])
    return ([], False, [])


def log_audit(conn, entity_type, entity_id, action, old_value=None, new_value=None, user_id=None):
    cursor = conn.cursor()
    cursor.execute(
        """INSERT INTO audit_log (entity_type, entity_id, action, old_value, new_value, user_id)
           VALUES (%s, %s, %s, %s, %s, %s)""",
        (entity_type, entity_id, action, old_value, new_value, user_id),
    )
    cursor.close()


PHT = ZoneInfo('Asia/Manila')

def to_pht(dt):
    """Convert a naive datetime to PHT (UTC+8) ISO string."""
    if dt is None:
        return None
    if isinstance(dt, str):
        return dt
    local_tz = datetime.now().astimezone().tzinfo
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=local_tz)
    return dt.astimezone(PHT).isoformat()


# ─── Auth / Login ─────────────────────────────────────────────────────────────

@app.route('/api/admin/login', methods=['POST'])
@limiter.limit("5 per minute")
def admin_login():
    data     = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()

    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            """SELECT id, name, email, role, branch, password, username, region, profile_pic, profile_zoom, profile_offset_y
               FROM team
               WHERE LOWER(TRIM(username)) = %s AND role = 'Admin'""",
            (normalize_username(username),)
        )
        row = cursor.fetchone()

        if not row:
            return jsonify({'error': 'Invalid credentials or account is not an Admin.'}), 401

        valid_password, should_rehash = verify_password(row[5], password)
        if not valid_password:
            return jsonify({'error': 'Invalid credentials or account is not an Admin.'}), 401

        if should_rehash:
            cursor.execute('UPDATE team SET password = %s WHERE id = %s', (generate_password_hash(password), row[0]))
            conn.commit()

        user = {
            'id':          row[0],
            'name':        row[1],
            'email':       row[2],
            'role':        row[3],
            'branch':      row[4],
            'username':    row[6],
            'region':      row[7],
            'profilePic':  row[8],
            'profileZoom': row[9],
            'profileOffsetY': row[10]
        }
        
        # Ensure identity is a string and include role/branch in claims
        access_token = create_access_token(
            identity=str(user['id']), 
            additional_claims={"role": user['role'], "branch": user['branch'], "region": user['region']}
        )
        return jsonify({'message': 'Login successful', 'user': user, 'access_token': access_token}), 200
    finally:
        close_connection(conn)


@app.route('/api/login', methods=['POST'])
@limiter.limit("5 per minute")
def login():
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '').strip()
    branch   = data.get('branch', '').strip()

    if not username or not password or not branch:
        return jsonify({'error': 'Username, password and branch are required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            """SELECT id, name, email, role, branch, password, username, region, profile_pic, profile_zoom, profile_offset_y
               FROM team
               WHERE LOWER(TRIM(username)) = %s
                 AND LOWER(TRIM(branch)) = %s""",
            (normalize_username(username), normalize_branch(branch))
        )
        row = cursor.fetchone()

        if not row:
            return jsonify({'error': 'Invalid credentials or branch mismatch'}), 401

        valid_password, should_rehash = verify_password(row[5], password)
        if not valid_password:
            return jsonify({'error': 'Invalid credentials or branch mismatch'}), 401

        if should_rehash:
            cursor.execute('UPDATE team SET password = %s WHERE id = %s', (generate_password_hash(password), row[0]))
            conn.commit()

        user = {
            'id':          row[0],
            'name':        row[1],
            'email':       row[2],
            'role':        row[3],
            'branch':      row[4],
            'username':    row[6],
            'region':      row[7],
            'profilePic':  row[8],
            'profileZoom': row[9],
            'profileOffsetY': row[10]
        }
        
        # Ensure identity is a string and include role/branch in claims
        access_token = create_access_token(
            identity=str(user['id']), 
            additional_claims={"role": user['role'], "branch": user['branch'], "region": user['region']}
        )
        return jsonify({'message': 'Login successful', 'user': user, 'access_token': access_token}), 200
    finally:
        close_connection(conn)


# ─── Customers ────────────────────────────────────────────────────────────────

@app.route('/api/customers', methods=['GET'])
@jwt_required()
def get_customers():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')

        # Base query for customers with computed status and KPIs
        # Status logic: Negotiation > Prospect > Converted > New
        query = """
            SELECT 
                c.id, c.name, c.industry, c.website, c.city, t.name AS owner, c.owner_id AS ownerId, 
                c.status AS companyStatus, c.created_at AS createdAt,
                l.contact_num AS contactNum, l.address, l.region, COALESCE(t_lead.name, l.owner_name, t.name) AS sr, l.branch,
                l.reassigned_at AS reassignedAt,
                COALESCE(deal_stats.totalDealCount, 0) AS totalDealCount,
                COALESCE(deal_stats.activeDealCount, 0) AS activeDealCount,
                COALESCE(deal_stats.closedWonCount, 0) AS closedWonCount,
                COALESCE(deal_stats.closedLostCount, 0) AS closedLostCount,
                COALESCE(deal_stats.closedWonValue, 0) AS closedWonValue,
                COALESCE(deal_stats.closedLostValue, 0) AS closedLostValue,
                CASE
                    WHEN COALESCE(deal_stats.closedLostCount, 0) = 0 THEN
                        CASE WHEN COALESCE(deal_stats.closedWonCount, 0) > 0 THEN 'Win Only' ELSE 'No History' END
                    ELSE ROUND(COALESCE(deal_stats.closedWonCount, 0) / COALESCE(deal_stats.closedLostCount, 0), 2)
                END AS winLossRatio,
                CASE
                    WHEN COALESCE(deal_stats.totalDealCount, 0) = 0 THEN 'New'
                    WHEN deal_stats.hasNegotiation > 0 THEN 'Negotiation'
                    WHEN deal_stats.hasProspect > 0 THEN 'Prospect'
                    WHEN deal_stats.activeDealCount = 0 AND deal_stats.totalDealCount > 0 THEN 'Converted'
                    ELSE 'New'
                END AS customerStatus
            FROM companies c
            LEFT JOIN team t ON c.owner_id = t.id
            LEFT JOIN leads l ON c.id = l.id
            LEFT JOIN team t_lead ON l.owner_id = t_lead.id
            LEFT JOIN (
                SELECT 
                    company_id,
                    COUNT(*) AS totalDealCount,
                    SUM(CASE WHEN stage NOT IN ('Closed Won', 'Closed Lost') THEN 1 ELSE 0 END) AS activeDealCount,
                    SUM(CASE WHEN stage = 'Closed Won' THEN 1 ELSE 0 END) AS closedWonCount,
                    SUM(CASE WHEN stage = 'Closed Lost' THEN 1 ELSE 0 END) AS closedLostCount,
                    SUM(CASE WHEN stage = 'Closed Won' THEN value ELSE 0 END) AS closedWonValue,
                    SUM(CASE WHEN stage = 'Closed Lost' THEN value ELSE 0 END) AS closedLostValue,
                    SUM(CASE WHEN stage IN ('Proposal', 'Negotiation') THEN 1 ELSE 0 END) AS hasNegotiation,
                    SUM(CASE WHEN stage IN ('New Opportunity', 'Qualified') THEN 1 ELSE 0 END) AS hasProspect
                FROM deals
                GROUP BY company_id
            ) deal_stats ON c.id = deal_stats.company_id
        """
        
        claims = get_jwt()
        user_id = get_jwt_identity()
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        where_clauses = ["1=1"] + list(scope_parts)
        params = list(scope_params)
        if restrict_owner:
            where_clauses.append("l.owner_id = %s")
            params.append(user_id)

        if where_clauses:
            query += " WHERE " + " AND ".join(where_clauses)

        query += " ORDER BY c.name"
        
        cursor.execute(query, tuple(params))
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/customers/<customer_id>', methods=['GET'])
@jwt_required()
def get_customer_detail(customer_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        claims = get_jwt()
        user_id = get_jwt_identity()
        
        # 1. Fetch Company/Customer with scoping
        scope_parts, restrict_owner, scope_params = build_scope(claims, '', col='LOWER(TRIM(l.branch))')
        where_parts = list(scope_parts) + ['c.id = %s']
        params = list(scope_params) + [customer_id]
        if restrict_owner:
            where_parts.append('l.owner_id = %s')
            params.append(user_id)
        
        where_sql = 'WHERE ' + ' AND '.join(where_parts)
        
        cursor.execute(f"""
            SELECT 
                c.id, c.name, c.industry, c.website, c.city, t.name AS owner, c.owner_id AS ownerId, 
                c.status AS companyStatus, c.created_at AS createdAt,
                l.contact_num AS contactNum, l.address, l.region, COALESCE(t_lead.name, l.owner_name, t.name) AS sr, l.branch,
                l.reassigned_at AS reassignedAt,
                COALESCE(deal_stats.totalDealCount, 0) AS totalDealCount,
                COALESCE(deal_stats.activeDealCount, 0) AS activeDealCount,
                COALESCE(deal_stats.closedWonCount, 0) AS closedWonCount,
                COALESCE(deal_stats.closedLostCount, 0) AS closedLostCount,
                COALESCE(deal_stats.closedWonValue, 0) AS closedWonValue,
                COALESCE(deal_stats.closedLostValue, 0) AS closedLostValue,
                CASE
                    WHEN COALESCE(deal_stats.totalDealCount, 0) = 0 THEN 'New'
                    WHEN deal_stats.hasNegotiation > 0 THEN 'Negotiation'
                    WHEN deal_stats.hasProspect > 0 THEN 'Prospect'
                    WHEN deal_stats.activeDealCount = 0 AND deal_stats.totalDealCount > 0 THEN 'Converted'
                    ELSE 'New'
                END AS customerStatus
            FROM companies c
            LEFT JOIN team t ON c.owner_id = t.id
            LEFT JOIN leads l ON c.id = l.id
            LEFT JOIN team t_lead ON l.owner_id = t_lead.id
            LEFT JOIN (
                SELECT 
                    company_id,
                    COUNT(*) AS totalDealCount,
                    SUM(CASE WHEN stage NOT IN ('Closed Won', 'Closed Lost') THEN 1 ELSE 0 END) AS activeDealCount,
                    SUM(CASE WHEN stage = 'Closed Won' THEN 1 ELSE 0 END) AS closedWonCount,
                    SUM(CASE WHEN stage = 'Closed Lost' THEN 1 ELSE 0 END) AS closedLostCount,
                    SUM(CASE WHEN stage = 'Closed Won' THEN value ELSE 0 END) AS closedWonValue,
                    SUM(CASE WHEN stage = 'Closed Lost' THEN value ELSE 0 END) AS closedLostValue,
                    SUM(CASE WHEN stage IN ('Proposal', 'Negotiation') THEN 1 ELSE 0 END) AS hasNegotiation,
                    SUM(CASE WHEN stage IN ('New Opportunity', 'Qualified') THEN 1 ELSE 0 END) AS hasProspect
                FROM deals
                GROUP BY company_id
            ) deal_stats ON c.id = deal_stats.company_id
            {where_sql}
        """, tuple(params))
        
        customer_row = cursor.fetchone()
        if not customer_row:
            return jsonify({'error': 'Customer not found or access denied'}), 404
            
        columns = [col[0] for col in cursor.description]
        customer_data = dict(zip(columns, customer_row))
        
        # 2. Fetch all deals for this customer
        deal_where = ["d.company_id = %s"]
        deal_params = [customer_id]
        if restrict_owner:
            deal_where.append("d.owner_id = %s")
            deal_params.append(user_id)
            
        cursor.execute(f"""
            SELECT d.id, d.name, d.stage, d.value, d.close_date AS closeDate, d.probability, 
                   t.name AS owner, d.owner_id AS ownerId, d.created_at AS createdAt
            FROM deals d
            LEFT JOIN team t ON d.owner_id = t.id
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            WHERE {" AND ".join(deal_where)}
            ORDER BY d.created_at DESC
        """, tuple(deal_params))
        deals = rows_to_list(cursor)
        
        # 3. Fetch all activities for these deals
        deal_ids = [d['id'] for d in deals]
        activities = []
        if deal_ids:
            format_strings = ','.join(['%s'] * len(deal_ids))
            cursor.execute(f"""
                SELECT a.id, a.subject, a.type, t.name AS owner, a.owner_id AS ownerId,
                       a.deal_id AS dealId, a.due_date AS dueDate, a.priority, a.status, a.notes,
                       a.created_at AS createdAt, a.stage
                FROM activities a
                LEFT JOIN team t ON a.owner_id = t.id
                WHERE a.deal_id IN ({format_strings})
                ORDER BY a.created_at DESC
            """, tuple(deal_ids))
            activities = rows_to_list(cursor)
            
        # 4. Fetch contacts for this customer
        cursor.execute(
            """SELECT c.id, c.name, c.role, c.email, c.phone, c.company_id AS companyId
               FROM contacts c
               WHERE c.company_id = %s
               ORDER BY c.name""",
            (customer_id,)
        )
        contacts = rows_to_list(cursor)

        # 5. Fetch audit log for these deals and contacts
        audit_logs = []
        if deal_ids or contacts:
            d_format = ','.join(['%s'] * len(deal_ids)) if deal_ids else "''"
            c_format = ','.join(['%s'] * len(contacts)) if contacts else "''"
            
            contact_ids = [c['id'] for c in contacts]
            params = deal_ids + contact_ids
            
            sql = f"""
                SELECT a.id, a.entity_type AS entityType, a.entity_id AS entityId, a.action, 
                       a.old_value AS oldValue, a.new_value AS newValue, a.changed_at AS changedAt,
                       t.name AS changedBy
                FROM audit_log a
                LEFT JOIN team t ON a.user_id = t.id
                WHERE (a.entity_type = 'deal' AND a.entity_id IN ({d_format}))
                   OR (a.entity_type = 'contact' AND a.entity_id IN ({c_format}))
                ORDER BY a.changed_at DESC
            """
            cursor.execute(sql, tuple(params))
            audit_logs = rows_to_list(cursor)
            for log in audit_logs:
                log['changedAt'] = to_pht(log.get('changedAt'))

        return jsonify({
            'customer': customer_data,
            'deals': deals,
            'activities': activities,
            'auditLogs': audit_logs,
            'contacts': contacts,
        })
    finally:
        close_connection(conn)


@app.route('/api/customers/<customer_id>/acknowledge', methods=['PATCH'])
@jwt_required()
def acknowledge_customer(customer_id):
    """Clear the 'newly assigned' flag once the assigned owner views the customer."""
    user_id = get_jwt_identity()
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            'UPDATE leads SET reassigned_at = NULL WHERE id = %s AND owner_id = %s',
            (customer_id, user_id)
        )
        conn.commit()
        return jsonify({'success': True})
    finally:
        close_connection(conn)


# ─── Team ─────────────────────────────────────────────────────────────────────

@app.route('/api/team', methods=['GET'])
@jwt_required()
def get_team():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        purpose = request.args.get('purpose', '')
        claims = get_jwt()
        role = claims.get('role', 'Branch Account')
        user_region = claims.get('region', '')
        user_branch = claims.get('branch', '')

        # Filter mode: return ALL team members in the caller's visibility scope,
        # regardless of rank (incl. self / peer managers). Used only to populate
        # view-filter dropdowns — NOT assignment targets, so the role hierarchy
        # restriction below is deliberately skipped. Default behavior is unchanged.
        if purpose == 'filter':
            if role in ('Branch Account', 'Sales Representative', 'Sales Rep'):
                cursor.execute(
                    'SELECT id, name, role, branch, region FROM team WHERE branch = %s ORDER BY name',
                    (user_branch,),
                )
            elif role == 'Regional Sales Manager':
                region_branches = REGION_BRANCHES.get(user_region, [])
                req = normalize_branch(branch)
                allowed_normalized = [normalize_branch(b) for b in region_branches]
                if req and req in allowed_normalized:
                    match = next((b for b in region_branches if normalize_branch(b) == req), None)
                    cursor.execute('SELECT id, name, role, branch, region FROM team WHERE branch = %s ORDER BY name', (match,))
                elif region_branches:
                    ph_branches = ', '.join(['%s'] * len(region_branches))
                    cursor.execute(
                        f'SELECT id, name, role, branch, region FROM team WHERE branch IN ({ph_branches}) ORDER BY name',
                        tuple(region_branches),
                    )
                else:
                    cursor.execute('SELECT id, name, role, branch, region FROM team WHERE 1=0')
            else:  # Head of Sales / Admin
                if branch and branch != 'Headquarters':
                    cursor.execute('SELECT id, name, role, branch, region FROM team WHERE branch = %s ORDER BY name', (branch,))
                else:
                    cursor.execute('SELECT id, name, role, branch, region FROM team ORDER BY name')
            return jsonify(rows_to_list(cursor))

        if role in ('Branch Account', 'Sales Representative', 'Sales Rep'):
            cursor.execute(
                'SELECT id, name, role, branch, region FROM team WHERE branch = %s AND role IN ("Sales Representative", "Branch Account", "Sales Rep") ORDER BY name',
                (user_branch,),
            )
        elif role == 'Regional Sales Manager':
            region_branches = REGION_BRANCHES.get(user_region, [])
            req = normalize_branch(branch)
            allowed_normalized = [normalize_branch(b) for b in region_branches]
            # RSM can only assign to SR
            target_roles = ('Sales Representative', 'Branch Account', 'Sales Rep')
            ph_roles = ', '.join(['%s'] * len(target_roles))
            if req and req in allowed_normalized:
                match = next((b for b in region_branches if normalize_branch(b) == req), None)
                cursor.execute(
                    f'SELECT id, name, role, branch, region FROM team WHERE branch = %s AND role IN ({ph_roles}) ORDER BY name',
                    (match, *target_roles),
                )
            elif region_branches:
                ph_branches = ', '.join(['%s'] * len(region_branches))
                cursor.execute(
                    f'SELECT id, name, role, branch, region FROM team WHERE branch IN ({ph_branches}) AND role IN ({ph_roles}) ORDER BY name',
                    tuple(region_branches + list(target_roles)),
                )
            else:
                cursor.execute('SELECT id, name, role, branch, region FROM team WHERE 1=0')
        elif role == 'Head of Sales':
            # HoS can assign to RSM and SR
            target_roles = ('Regional Sales Manager', 'Sales Representative', 'Branch Account', 'Sales Rep')
            ph_roles = ', '.join(['%s'] * len(target_roles))
            if branch and branch != 'Headquarters':
                cursor.execute(
                    f'SELECT id, name, role, branch, region FROM team WHERE branch = %s AND role IN ({ph_roles}) ORDER BY name',
                    (branch, *target_roles),
                )
            else:
                cursor.execute(f'SELECT id, name, role, branch, region FROM team WHERE role IN ({ph_roles}) ORDER BY name', target_roles)
        else:
            if branch and branch != 'Headquarters':
                cursor.execute(
                    'SELECT id, name, role, branch, region FROM team WHERE branch = %s ORDER BY name',
                    (branch,),
                )
            else:
                cursor.execute('SELECT id, name, role, branch, region FROM team ORDER BY name')

        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


# ─── Companies ────────────────────────────────────────────────────────────────

@app.route('/api/companies', methods=['GET'])
@jwt_required()
def get_companies():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        if restrict_owner:
            # SR: include companies they own directly OR companies referenced by their deals.
            # The second condition covers Agos-synced companies where owner_id was never set
            # because the sync predates the cascade fix.
            where_sql = ('WHERE (c.owner_id = %s '
                         'OR c.id IN (SELECT d.company_id FROM deals d WHERE d.owner_id = %s AND d.company_id IS NOT NULL))')
            params_tuple = (user_id, user_id)
        else:
            where_parts = list(scope_parts)
            where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''
            params_tuple = tuple(scope_params)
        cursor.execute(f'''
            SELECT c.id, c.name, c.industry, c.website, c.city, t.name AS owner, c.owner_id AS ownerId, c.status, c.created_at
            FROM companies c
            LEFT JOIN team t ON c.owner_id = t.id
            LEFT JOIN leads l ON l.id = c.id
            {where_sql}
            ORDER BY c.name
        ''', params_tuple)
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/companies', methods=['POST'])
@jwt_required()
def create_company():
    data = request.get_json()
    required = ['id', 'name']
    if not all(k in data for k in required):
        return jsonify({'error': 'id and name are required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        user_id = get_jwt_identity()
        claims = get_jwt()
        
        # Default owner and branch from current user if not provided
        owner_id = data.get('ownerId') or user_id
        
        # Fetch branch from team if not provided
        branch = data.get('branch')
        region = data.get('region')
        if not branch:
            cursor.execute('SELECT branch, region FROM team WHERE id = %s', (user_id,))
            user_row = cursor.fetchone()
            if user_row:
                branch = user_row[0]
                region = region or user_row[1]

        # 1. Insert into leads (Master record)
        cursor.execute(
            """INSERT INTO leads (id, customer_name, branch, region, owner_id, status)
               VALUES (%s, %s, %s, %s, %s, %s)
               ON DUPLICATE KEY UPDATE customer_name = VALUES(customer_name)""",
            (data['id'], data['name'], branch, region, owner_id, 'New')
        )

        # 2. Insert into companies
        cursor.execute(
            """INSERT INTO companies (id, name, industry, website, city, owner_id, status)
               VALUES (%s, %s, %s, %s, %s, %s, %s)
               ON DUPLICATE KEY UPDATE name = VALUES(name), owner_id = VALUES(owner_id)""",
            (
                data['id'],
                data['name'],
                data.get('industry'),
                data.get('website'),
                data.get('city'),
                owner_id,
                data.get('status', 'Active'),
            ),
        )
        conn.commit()
        return jsonify({
            'id': data['id'],
            'name': data['name'],
            'industry': data.get('industry'),
            'website': data.get('website'),
            'city': data.get('city'),
            'ownerId': owner_id,
            'branch': branch,
            'status': data.get('status', 'Active')
        }), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


# ─── Contacts ─────────────────────────────────────────────────────────────────

@app.route('/api/contacts', methods=['GET'])
@jwt_required()
def get_contacts():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        where_parts = list(scope_parts)
        params = list(scope_params)
        if restrict_owner:
            where_parts.append('l.owner_id = %s')
            params.append(user_id)
        where_sql = 'WHERE ' + ' AND '.join(where_parts) if where_parts else ''
        cursor.execute(f'''
            SELECT c.id, c.name, c.company_id AS companyId, c.role, t.name AS owner, c.owner_id AS ownerId,
                   c.email, c.phone, c.last_touch AS lastTouch, c.status, c.created_at
            FROM contacts c
            LEFT JOIN team t ON c.owner_id = t.id
            LEFT JOIN leads l ON l.id = c.company_id
            {where_sql}
            ORDER BY c.name
        ''', tuple(params))
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/contacts', methods=['POST'])
@jwt_required()
def create_contact():
    data = request.get_json()
    if not all(k in data for k in ['id', 'name']):
        return jsonify({'error': 'id and name are required'}), 400
    if not data.get('email') and not data.get('phone'):
        return jsonify({'error': 'email or phone is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        user_id = get_jwt_identity()
        owner_id = data.get('ownerId') or user_id
        
        # Strict Duplicate Check: Name + (Email OR Phone)
        name = data['name'].strip()
        email = data.get('email', '').strip() or None
        phone = data.get('phone', '').strip() or None
        
        check_sql = "SELECT id, name FROM contacts WHERE LOWER(TRIM(name)) = %s AND ("
        check_params = [name.lower()]
        
        conditions = []
        if email:
            conditions.append("email = %s")
            check_params.append(email)
        if phone:
            conditions.append("phone = %s")
            check_params.append(phone)
            
        if not conditions:
            # If neither email nor phone provided (already handled by validation but safe to check)
            pass
        else:
            check_sql += " OR ".join(conditions) + ")"
            cursor.execute(check_sql, tuple(check_params))
            existing = cursor.fetchone()
            if existing:
                return jsonify({'error': f'A contact named "{existing[1]}" already exists with this email/phone.'}), 409

        # Insert into contacts with UPSERT capability
        cursor.execute(
            """INSERT INTO contacts (id, name, company_id, role, owner_id, email, phone, last_touch, status, created_at)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
               ON DUPLICATE KEY UPDATE 
                name = VALUES(name), 
                company_id = VALUES(company_id), 
                role = VALUES(role),
                email = VALUES(email),
                phone = VALUES(phone),
                status = VALUES(status)""",
            (
                data['id'],
                data['name'],
                data.get('companyId'),
                data.get('role'),
                owner_id,
                data.get('email'),
                data.get('phone'),
                data.get('lastTouch') or None,
                data.get('status', 'Active'),
                date.today()
            ),
        )
        
        # Optional: Sync primary contact number back to lead if lead's number is missing
        if data.get('companyId') and data.get('phone'):
            # Find the customer name to apply this to "all accounts" (all leads with the same name)
            cursor.execute("SELECT customer_name FROM leads WHERE id = %s", (data['companyId'],))
            name_row = cursor.fetchone()
            if name_row:
                cust_name = name_row[0]
                cursor.execute(
                    "UPDATE leads SET contact_num = %s WHERE customer_name = %s AND (contact_num IS NULL OR contact_num = '')",
                    (data['phone'], cust_name)
                )
        
        log_audit(conn, 'contact', data['id'], 'contact_created', None, data['name'], user_id)
        conn.commit()

        return jsonify({
            'id': data['id'],
            'name': data['name'],
            'companyId': data.get('companyId'),
            'role': data.get('role'),
            'ownerId': owner_id,
            'email': data.get('email'),
            'phone': data.get('phone'),
            'status': data.get('status', 'Active'),
            'createdAt': date.today().isoformat(),
            'lastTouch': data.get('lastTouch')
        }), 201
    finally:
        close_connection(conn)


@app.route('/api/contacts/<contact_id>', methods=['PUT'])
@jwt_required()
def update_contact(contact_id):
    data = request.get_json()
    name = data.get('name')
    role = data.get('role')
    email = data.get('email')
    phone = data.get('phone')

    if not name:
        return jsonify({'error': 'name is required'}), 400
    if not email and not phone:
        return jsonify({'error': 'email or phone is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        cursor.execute("SELECT name, role, email, phone FROM contacts WHERE id = %s", (contact_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Contact not found'}), 404

        old_name, old_role, old_email, old_phone = row

        # Duplicate check if name/email/phone changed
        if name != old_name or email != old_email or phone != old_phone:
            check_name = (name or old_name).strip().lower()
            check_email = (email or old_email).strip() or None
            check_phone = (phone or old_phone).strip() or None

            check_sql = "SELECT id, name FROM contacts WHERE id != %s AND LOWER(TRIM(name)) = %s AND ("
            check_params = [contact_id, check_name]
            conditions = []
            if check_email:
                conditions.append("email = %s")
                check_params.append(check_email)
            if check_phone:
                conditions.append("phone = %s")
                check_params.append(check_phone)
            
            if conditions:
                check_sql += " OR ".join(conditions) + ")"
                cursor.execute(check_sql, tuple(check_params))
                existing = cursor.fetchone()
                if existing:
                    return jsonify({'error': f'A contact named "{existing[1]}" already exists with this email/phone.'}), 409

        updates = []
        params = []
        if name != old_name:
            updates.append('name = %s')
            params.append(name)
        if role != old_role:
            updates.append('role = %s')
            params.append(role)
        if email != old_email:
            updates.append('email = %s')
            params.append(email)
        if phone != old_phone:
            updates.append('phone = %s')
            params.append(phone)

        if updates:
            params.append(contact_id)
            cursor.execute(f'UPDATE contacts SET {", ".join(updates)} WHERE id = %s', params)
            log_audit(conn, 'contact', contact_id, 'update',
                      json.dumps({'name': old_name, 'role': old_role, 'email': old_email, 'phone': old_phone}),
                      json.dumps({'name': name, 'role': role, 'email': email, 'phone': phone}),
                      get_jwt_identity())
            conn.commit()

        return jsonify({'message': 'Contact updated', 'id': contact_id})
    finally:
        close_connection(conn)


# ─── Google Sheets Sync ───────────────────────────────────────────────────────

@app.route('/api/sync/gsheets', methods=['POST'])
@limiter.limit("2 per minute")
@jwt_required()
@admin_required
def trigger_gsheets_sync():
    try:
        result = sync_from_sheets()
        if "error" in result:
            return jsonify(result), 500
        return jsonify(result), 200
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ─── Leads ────────────────────────────────────────────────────────────────────

@app.route('/api/leads', methods=['GET'])
@jwt_required()
def get_leads():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        where_parts = list(scope_parts)
        params = list(scope_params)
        if restrict_owner:
            where_parts.append('l.owner_id = %s')
            params.append(user_id)
        where_sql = 'WHERE ' + ' AND '.join(where_parts) if where_parts else ''



        cursor.execute(f'''
            SELECT l.id, l.customer_name AS customerName, l.contact_num AS contactNum,
                   l.address, l.region, t.name AS sr, l.owner_id AS ownerId, l.branch, l.status, l.created_at AS createdAt
            FROM leads l
            LEFT JOIN team t ON l.owner_id = t.id
            {where_sql}
            ORDER BY l.created_at DESC
        ''', tuple(params))
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/leads', methods=['POST'])
@jwt_required()
def create_lead():
    data = request.get_json()
    if not data.get('customerName'):
        return jsonify({'error': 'customerName is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        lead_id = data.get('id')
        customer_name = data.get('customerName')
        contact_num = data.get('contactNum')
        branch = data.get('branch')
        owner_id = data.get('ownerId')

        # Validation: Ensure owner_id belongs to the correct branch/region
        if owner_id:
            cursor.execute('SELECT branch, region, role FROM team WHERE id = %s', (owner_id,))
            team_row = cursor.fetchone()
            if not team_row:
                return jsonify({'error': 'Assigned owner not found in team'}), 400
            
            token_claims = get_jwt()
            token_role = token_claims.get('role', '')
            user_id = get_jwt_identity()

            # Role hierarchy check
            if not can_assign(token_claims, user_id, owner_id, team_row[2]):
                return jsonify({'error': f'Cannot assign lead to a user with equal or higher role ({team_row[2]})'}), 403

            if token_role not in ('Admin', 'Head of Sales'):
                if token_role == 'Regional Sales Manager':
                    rsm_region = token_claims.get('region', '')
                    allowed = [normalize_branch(b) for b in REGION_BRANCHES.get(rsm_region, [])]
                    if normalize_branch(branch) not in allowed:
                        return jsonify({'error': f'Branch ({branch}) is outside your region scope'}), 403
                elif normalize_branch(team_row[0]) != normalize_branch(branch):
                    return jsonify({'error': f'Owner branch ({team_row[0]}) mismatch with lead branch ({branch})'}), 400
        
        # 1. Insert into leads (Master record matching GSheets)
        cursor.execute(
            """INSERT INTO leads (id, customer_name, contact_num, address, region, owner_id, branch, status, created_at)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                lead_id,
                customer_name,
                contact_num,
                data.get('address'),
                data.get('region'),
                owner_id,
                branch,
                data.get('status', 'New'),
                data.get('createdAt') or date.today(),
            ),
        )
        
        # 2. Automatically create a Company record
        # Use lead_id as company_id for direct linking
        cursor.execute(
            """INSERT INTO companies (id, name, city, owner_id, status)
               VALUES (%s, %s, %s, %s, %s)
               ON DUPLICATE KEY UPDATE name = VALUES(name), owner_id = VALUES(owner_id)""",
            (lead_id, customer_name, data.get('region'), owner_id, 'Active')
        )
        
        # 3. Automatically create a Contact record (with duplicate prevention)
        import uuid
        contact_id = str(uuid.uuid4())
        
        # Global duplicate check: Name + Phone
        contact_exists = False
        if contact_num:
            cursor.execute(
                "SELECT id FROM contacts WHERE LOWER(TRIM(name)) = %s AND TRIM(phone) = %s",
                (customer_name.lower().strip(), contact_num.strip())
            )
            if cursor.fetchone():
                contact_exists = True

        if not contact_exists:
            cursor.execute(
                """INSERT INTO contacts (id, name, company_id, owner_id, phone, status)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (contact_id, customer_name, lead_id, owner_id, contact_num, 'Active')
            )
        
        # 4. Add the contact to the deal if it's created via pipeline sync (not applicable here, but good for consistency)
        # For now, we just ensure the join table exists for many-to-many.
        
        conn.commit()
        
        return jsonify({'message': 'Lead, Company, and Contact created', 'id': lead_id}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


@app.route('/api/leads/<lead_id>/status', methods=['PATCH'])
@jwt_required()
def update_lead_status(lead_id):
    data = request.get_json()
    new_status = data.get('status')
    if not new_status:
        return jsonify({'error': 'status is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT status FROM leads WHERE id = %s', (lead_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Lead not found'}), 404

        old_status = row[0]
        cursor.execute('UPDATE leads SET status = %s WHERE id = %s', (new_status, lead_id))
        log_audit(conn, 'lead', lead_id, 'status_change', old_status, new_status, get_jwt_identity())
        conn.commit()
        return jsonify({'message': 'Lead status updated'})
    finally:
        close_connection(conn)


@app.route('/api/leads/<lead_id>/reassign', methods=['PATCH'])
@jwt_required()
def reassign_lead(lead_id):
    claims = get_jwt()
    role = claims.get('role', '')
    if role not in ('Head of Sales', 'Regional Sales Manager'):
        return jsonify({'error': 'Insufficient permissions to reassign leads'}), 403

    data = request.get_json()
    new_owner_id = data.get('newOwnerId')
    if not new_owner_id:
        return jsonify({'error': 'newOwnerId is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT owner_id, branch FROM leads WHERE id = %s', (lead_id,))
        lead_row = cursor.fetchone()
        if not lead_row:
            return jsonify({'error': 'Lead not found'}), 404
        old_owner_id, lead_branch = lead_row

        cursor.execute('SELECT branch, region, role FROM team WHERE id = %s', (new_owner_id,))
        new_owner_row = cursor.fetchone()
        if not new_owner_row:
            return jsonify({'error': 'New owner not found in team'}), 400
        new_owner_branch, new_owner_region, new_owner_role = new_owner_row

        # Role hierarchy check
        user_id = get_jwt_identity()
        if not can_assign(claims, user_id, new_owner_id, new_owner_role):
            return jsonify({'error': f'Cannot reassign to a user with equal or higher role ({new_owner_role})'}), 403

        # RSM can only reassign within their region
        if role == 'Regional Sales Manager':
            user_region = claims.get('region', '')
            allowed = [normalize_branch(b) for b in REGION_BRANCHES.get(user_region, [])]
            if normalize_branch(new_owner_branch) not in allowed:
                return jsonify({'error': f'New owner branch ({new_owner_branch}) is outside your region'}), 403

        cursor.execute('UPDATE leads SET owner_id = %s, reassigned_at = CURRENT_DATE WHERE id = %s', (new_owner_id, lead_id))
        cursor.execute('UPDATE companies SET owner_id = %s WHERE id = %s', (new_owner_id, lead_id))
        log_audit(conn, 'lead', lead_id, 'reassign', str(old_owner_id), str(new_owner_id), get_jwt_identity())
        conn.commit()
        return jsonify({'message': 'Lead reassigned successfully'})
    finally:
        close_connection(conn)


# ─── Deals ────────────────────────────────────────────────────────────────────

@app.route('/api/deals', methods=['GET'])
@jwt_required()
def get_deals():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        where_parts = list(scope_parts) + ['l.branch IS NOT NULL']
        params = list(scope_params)
        if restrict_owner:
            where_parts.append('d.owner_id = %s')
            params.append(user_id)
        where_sql = 'WHERE ' + ' AND '.join(where_parts)
        cursor.execute(f'''
            SELECT d.id, d.name, d.company_id AS companyId, d.contact_id AS contactId,
                   d.lead_id AS leadId, d.stage, d.value, d.close_date AS closeDate,
                   d.probability, t.name AS owner, d.owner_id AS ownerId, t.role AS ownerRole, d.created_at,
                   d.lost_reason AS lostReason, l.branch AS branch,
                   COALESCE(urgency.urgencyScore, 0) AS urgencyScore,
                   CASE
                       WHEN urgency.hasOverdue = 1 THEN 'Overdue'
                       WHEN urgency.hasHigh = 1 THEN 'High Priority'
                       WHEN urgency.hasToday = 1 THEN 'Due Today'
                       ELSE NULL
                   END AS urgencyLabel,
                   urgency.nextDueDate,
                   GREATEST(
                       IFNULL(activity_touch.lastActivityDate, DATE('1000-01-01')),
                       IFNULL(DATE(audit_touch.lastAuditDate), DATE('1000-01-01')),
                       IFNULL(d.created_at, DATE('1000-01-01'))
                   ) AS lastTouch,
                   CASE WHEN agos.agosCount > 0 THEN 1 ELSE 0 END AS isAgos
            FROM deals d
            LEFT JOIN team t ON d.owner_id = t.id
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            LEFT JOIN (
                SELECT a.deal_id,
                       MAX(CASE
                           WHEN a.status IN ('Open', 'Reopened') AND a.due_date < CURDATE() THEN 3
                           WHEN a.status IN ('Open', 'Reopened') AND a.priority = 'High' THEN 2
                           WHEN a.status IN ('Open', 'Reopened') AND a.due_date = CURDATE() THEN 1
                           ELSE 0
                       END) AS urgencyScore,
                       MAX(CASE WHEN a.status IN ('Open', 'Reopened') AND a.due_date < CURDATE() THEN 1 ELSE 0 END) AS hasOverdue,
                       MAX(CASE WHEN a.status IN ('Open', 'Reopened') AND a.priority = 'High' THEN 1 ELSE 0 END) AS hasHigh,
                       MAX(CASE WHEN a.status IN ('Open', 'Reopened') AND a.due_date = CURDATE() THEN 1 ELSE 0 END) AS hasToday,
                       MIN(CASE WHEN a.status IN ('Open', 'Reopened') THEN a.due_date END) AS nextDueDate
                FROM activities a
                GROUP BY a.deal_id
            ) urgency ON urgency.deal_id = d.id
            LEFT JOIN (
                SELECT a.deal_id, MAX(a.created_at) AS lastActivityDate
                FROM activities a
                GROUP BY a.deal_id
            ) activity_touch ON activity_touch.deal_id = d.id
            LEFT JOIN (
                SELECT al.entity_id AS deal_id, MAX(al.changed_at) AS lastAuditDate
                FROM audit_log al
                WHERE al.entity_type = 'deal'
                GROUP BY al.entity_id
            ) audit_touch ON audit_touch.deal_id = d.id
            LEFT JOIN (
                SELECT a.deal_id,
                       SUM(CASE WHEN a.notes LIKE '%Automatic task generated%' THEN 1 ELSE 0 END) AS agosCount
                FROM activities a
                GROUP BY a.deal_id
            ) agos ON agos.deal_id = d.id
            {where_sql}
            ORDER BY urgencyScore DESC, d.created_at DESC
        ''', tuple(params))
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/deals', methods=['POST'])
@jwt_required()
def create_deal():
    data = request.get_json()
    if not all(k in data for k in ['id', 'name']):
        return jsonify({'error': 'id and name are required'}), 400

    stage = data.get('stage', 'Qualified')
    probability = STAGE_PROBABILITY.get(stage, 20)

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        # Check for existing deal with same name for this company
        deal_name = data.get('name', '').strip()
        company_id = data.get('companyId')
        if deal_name and company_id:
            cursor.execute(
                'SELECT id FROM deals WHERE company_id = %s AND LOWER(TRIM(name)) = %s',
                (company_id, deal_name.lower())
            )
            if cursor.fetchone():
                return jsonify({'error': f'A deal named "{deal_name}" already exists for this company.'}), 409

        owner_id = data.get('ownerId')
        # Validation for owner_id branch mismatch
        if owner_id:
            cursor.execute('SELECT branch, region, role FROM team WHERE id = %s', (owner_id,))
            team_row = cursor.fetchone()
            if not team_row:
                return jsonify({'error': 'Assigned owner not found in team'}), 400

            token_claims = get_jwt()
            user_id = get_jwt_identity()

            # Role hierarchy check
            if not can_assign(token_claims, user_id, owner_id, team_row[2]):
                return jsonify({'error': f'Cannot assign deal to a user with equal or higher role ({team_row[2]})'}), 403

            # If leadId is provided, check branch against lead's branch
            lead_id = data.get('leadId')
            if lead_id:
                cursor.execute('SELECT branch FROM leads WHERE id = %s', (lead_id,))
                lead_row = cursor.fetchone()
                if lead_row:
                    if team_row[0] != lead_row[0]:
                        token_claims = get_jwt()
                        token_role = token_claims.get('role', '')
                        if token_role not in ('Admin', 'Head of Sales'):
                            if token_role == 'Regional Sales Manager':
                                rsm_region = token_claims.get('region', '')
                                allowed = [normalize_branch(b) for b in REGION_BRANCHES.get(rsm_region, [])]
                                if normalize_branch(lead_row[0]) not in allowed:
                                    return jsonify({'error': f'Lead branch ({lead_row[0]}) is outside your region scope'}), 403
                            else:
                                return jsonify({'error': f'Owner branch ({team_row[0]}) mismatch with lead/deal branch ({lead_row[0]})'}), 400

        cursor.execute(
            """INSERT INTO deals (id, name, company_id, contact_id, lead_id, stage, value, close_date, probability, owner_id)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                data['id'],
                data['name'],
                data.get('companyId'),
                data.get('contactId'),
                data.get('leadId') or None,
                stage,
                data.get('value', 0),
                data.get('closeDate') or data.get('expectedClose') or None,
                probability,
                data.get('ownerId'),
            ),
        )
        log_audit(conn, 'deal', data['id'], 'deal_created', None, stage, get_jwt_identity())
        if data.get('ownerId') and data.get('companyId'):
            cursor.execute(
                'UPDATE leads SET owner_id = %s, reassigned_at = CURRENT_DATE WHERE id = %s AND (owner_id IS NULL OR owner_id != %s)',
                (data['ownerId'], data['companyId'], data['ownerId'])
            )
            cursor.execute(
                'UPDATE companies SET owner_id = %s WHERE id = %s',
                (data['ownerId'], data['companyId'])
            )
        conn.commit()
        return jsonify({'message': 'Deal created', 'id': data['id']}), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


@app.route('/api/deals/<deal_id>/stage', methods=['PATCH'])
@jwt_required()
def update_deal_stage(deal_id):
    data = request.get_json()
    new_stage = data.get('stage')
    new_value = data.get('value')
    new_close = data.get('closeDate')
    new_owner = data.get('owner')
    new_name = data.get('name')

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT stage, value, close_date, owner_id, lead_id, probability, probability_manual, lost_reason, name, company_id FROM deals WHERE id = %s', (deal_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Deal not found'}), 404

        old_stage, old_value, old_close, old_owner_id, lead_id, old_probability, old_probability_manual, old_lost_reason, old_name, company_id = row

        claims = get_jwt()
        user_id = get_jwt_identity()
        user_role = claims.get('role', '')
        if str(old_owner_id) != str(user_id):
            if ROLE_RANK.get(user_role, 0) <= 40:
                # SR / Branch Account may only touch their own deals
                return jsonify({'error': 'Only the assigned SR can update this deal'}), 403
            if user_role != 'Admin':
                # Managers may edit deals owned by SRs, but not by another manager
                cursor.execute('SELECT role FROM team WHERE id = %s', (old_owner_id,))
                owner_role_row = cursor.fetchone()
                owner_rank = ROLE_RANK.get(owner_role_row[0], 0) if owner_role_row else 0
                if owner_rank > 40:
                    return jsonify({'error': 'You cannot edit a deal owned by another manager'}), 403

        updates = []
        params = []

        if new_name and new_name.strip().lower() != old_name.strip().lower():
            # Check for duplicates
            cursor.execute(
                'SELECT id FROM deals WHERE company_id = %s AND LOWER(TRIM(name)) = %s AND id != %s',
                (company_id, new_name.strip().lower(), deal_id)
            )
            if cursor.fetchone():
                return jsonify({'error': f'A deal named "{new_name}" already exists for this company.'}), 409
            
            updates.append('name = %s')
            params.append(new_name.strip())
            log_audit(conn, 'deal', deal_id, 'name_change', old_name, new_name.strip(), get_jwt_identity())
        
        probability = data.get('probability')
        if probability is not None and probability != old_probability:
            updates.append('probability = %s')
            params.append(probability)
            updates.append('probability_manual = TRUE')
            log_audit(conn, 'deal', deal_id, 'probability_change', str(old_probability), str(probability), get_jwt_identity())
        
        if new_stage and new_stage != old_stage:
            if probability is not None:
                # Both probability and stage sent — stage updated, probability already set above
                updates.append('stage = %s')
                params.append(new_stage)
            elif old_probability_manual:
                # SR manually set probability before — preserve it
                updates.append('stage = %s')
                params.append(new_stage)
            else:
                # Auto-assign probability from stage
                new_probability = STAGE_PROBABILITY.get(new_stage, 20)
                updates.append('stage = %s, probability = %s')
                params.extend([new_stage, new_probability])
                updates.append('probability_manual = FALSE')
            log_audit(conn, 'deal', deal_id, 'stage_change', old_stage, new_stage, get_jwt_identity())
        
        if new_value is not None and float(new_value) != float(old_value):
            updates.append('value = %s')
            params.append(new_value)
            log_audit(conn, 'deal', deal_id, 'value_change', str(old_value), str(new_value), get_jwt_identity())
            
        if new_close and str(new_close) != str(old_close):
            updates.append('close_date = %s')
            params.append(new_close)
            log_audit(conn, 'deal', deal_id, 'close_date_change', str(old_close), str(new_close), get_jwt_identity())

        if data.get('ownerId'):
            new_owner_id = data.get('ownerId')
            if str(new_owner_id) != str(old_owner_id):
                cursor.execute('SELECT name, role, branch FROM team WHERE id = %s', (new_owner_id,))
                target_row = cursor.fetchone()
                if target_row and not can_assign(claims, user_id, new_owner_id, target_row[1]):
                    return jsonify({'error': f'Cannot assign deal to a user with equal or higher role ({target_row[1]})'}), 403

                # Enforce same-branch reassignment
                cursor.execute('SELECT branch FROM leads WHERE id = %s', (lead_id,))
                deal_branch_row = cursor.fetchone()
                deal_branch = (deal_branch_row[0] or '').lower()
                target_branch = (target_row[2] or '').lower() if target_row else ''
                if deal_branch and target_branch and deal_branch != target_branch:
                    return jsonify({'error': 'Cannot reassign deal to an SR from a different branch'}), 403

                updates.append('owner_id = %s')
                params.append(new_owner_id)
                log_audit(conn, 'deal', deal_id, 'owner_id_change', str(old_owner_id), str(new_owner_id), get_jwt_identity())
                cursor.execute(
                    'UPDATE leads SET owner_id = %s, reassigned_at = CURRENT_DATE WHERE id = %s AND (owner_id IS NULL OR owner_id != %s)',
                    (new_owner_id, company_id, new_owner_id)
                )
                cursor.execute(
                    'UPDATE companies SET owner_id = %s WHERE id = %s',
                    (new_owner_id, company_id)
                )

        lost_reason = data.get('lostReason')
        if new_stage == 'Closed Lost' and lost_reason and lost_reason != old_lost_reason:
            updates.append('lost_reason = %s')
            params.append(lost_reason)
            log_audit(conn, 'deal', deal_id, 'lost_reason', old_lost_reason, lost_reason, get_jwt_identity())

        if not updates:
            return jsonify({'message': 'No updates provided'}), 400

        params.append(deal_id)
        cursor.execute(f'UPDATE deals SET {", ".join(updates)} WHERE id = %s', params)

        # Automatically close all open tasks if deal is closed
        if new_stage in ('Closed Won', 'Closed Lost'):
            cursor.execute(
                "UPDATE activities SET status = 'Completed' WHERE deal_id = %s AND status IN ('Open', 'Reopened')",
                (deal_id,)
            )
            log_audit(conn, 'deal', deal_id, 'bulk_task_completion', 'Open/Reopened', 'Completed', user_id)

        conn.commit()

        # Log activity for edit-detail changes (value, closeDate, probability)
        editor_name = 'Unknown'
        cursor.execute('SELECT name FROM team WHERE id = %s', (user_id,))
        user_row = cursor.fetchone()
        if user_row:
            editor_name = user_row[0]

        today_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        change_desc = []
        if new_value is not None and float(new_value) != float(old_value):
            change_desc.append(f'value to {new_value}')
        if new_close and str(new_close) != str(old_close):
            change_desc.append(f'close date to {new_close}')
        if probability is not None and probability != old_probability:
            change_desc.append(f'probability to {probability}%')

        new_activity = None
        if change_desc:
            activity_id = f'update-{deal_id}-{int(datetime.now().timestamp())}'
            if len(change_desc) == 1:
                subject = f'{editor_name} updated deal {change_desc[0]}'
            else:
                subject = f'{editor_name} updated: {", ".join(change_desc)}'
            cursor.execute(
                """INSERT INTO activities (id, subject, type, deal_id, status, created_at)
                   VALUES (%s, %s, %s, %s, %s, %s)""",
                (activity_id, subject, 'Update', deal_id, 'Completed', today_str),
            )
            conn.commit()
            new_activity = {
                'id': activity_id,
                'subject': subject,
                'title': subject,
                'type': 'Update',
                'dealId': deal_id,
                'status': 'Completed',
                'dueDate': None,
                'priority': 'Medium',
                'companyName': '',
                'contact': '',
                'notes': None,
                'created_at': today_str,
            }

        if new_stage:
            is_closed = new_stage in ('Closed Won', 'Closed Lost')
            if is_closed and lead_id:
                cursor.execute('SELECT branch FROM leads WHERE id = %s', (lead_id,))
                lead_row = cursor.fetchone()
                if lead_row:
                    fill_pipeline(conn, branch=lead_row[0])
                    conn.commit()

        result = {'message': 'Deal updated successfully'}
        if new_activity:
            result['activity'] = new_activity
        return jsonify(result)
    finally:
        close_connection(conn)


# ─── Activities (Tasks) ───────────────────────────────────────────────────────

@app.route('/api/activities', methods=['GET'])
@jwt_required()
def get_activities():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        where_parts = list(scope_parts)
        params = list(scope_params)
        if restrict_owner:
            where_parts.append('a.owner_id = %s')
            params.append(user_id)
        where_sql = 'WHERE ' + ' AND '.join(where_parts) if where_parts else 'WHERE 1=1'
        cursor.execute(f'''
            SELECT a.id, a.subject, a.type, t.name AS owner, a.owner_id AS ownerId, a.deal_id AS dealId,
                   a.due_date AS dueDate, a.priority, a.status, a.notes, a.created_at, a.stage,
                   a.contact_name, d.name AS dealName,
                   COALESCE(c.name, l.customer_name, '') AS companyName
            FROM activities a
            LEFT JOIN team t ON a.owner_id = t.id
            LEFT JOIN deals d ON d.id = a.deal_id
            LEFT JOIN companies c ON d.company_id = c.id
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            {where_sql}
            ORDER BY
                CASE
                    WHEN a.status IN ('Open', 'Reopened') AND a.due_date < CURDATE() THEN 1
                    WHEN a.status IN ('Open', 'Reopened') AND a.priority = 'High' THEN 2
                    WHEN a.status IN ('Open', 'Reopened') AND a.due_date = CURDATE() THEN 3
                    WHEN a.status IN ('Open', 'Reopened') THEN 4
                    ELSE 5
                END,
                a.due_date ASC
        ''', tuple(params))
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/activities', methods=['POST'])
@jwt_required()
def create_activity():
    data = request.get_json()
    if not all(k in data for k in ['id', 'subject']):
        return jsonify({'error': 'id and subject are required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        # Security: Resolve owner ID by name if provided, instead of trusting raw ID
        # This prevents vulnerability of frontend manipulating database IDs
        final_owner_id = data.get('ownerId')
        owner_name = data.get('owner') # Frontend sends name string here
        
        if owner_name:
            cursor.execute("SELECT id FROM team WHERE name = %s", (owner_name,))
            row = cursor.fetchone()
            if row:
                final_owner_id = row[0]

        cursor.execute(
            """INSERT INTO activities (id, subject, type, owner_id, deal_id, due_date, priority, status, notes, stage, contact_name, metadata)
               VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
            (
                data['id'],
                data['subject'],
                data.get('type', 'Follow-up'),
                final_owner_id,
                data.get('dealId') or None,
                data.get('dueDate') or None,
                data.get('priority', 'Medium'),
                data.get('status', 'Open'),
                data.get('notes'),
                data.get('stage'),
                data.get('contact'),
                data.get('metadata'),
            ),
        )
        conn.commit()
        return jsonify({'message': 'Activity created', 'id': data['id']}), 201
    finally:
        close_connection(conn)


@app.route('/api/activities/<activity_id>/status', methods=['PATCH'])
@jwt_required()
def update_activity_status(activity_id):
    data = request.get_json()
    new_status = data.get('status')
    if not new_status:
        return jsonify({'error': 'status is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT status FROM activities WHERE id = %s', (activity_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Activity not found'}), 404

        old_status = row[0]
        if new_status != old_status:
            cursor.execute('UPDATE activities SET status = %s WHERE id = %s', (new_status, activity_id))
            log_audit(conn, 'activity', activity_id, 'status_change', old_status, new_status, get_jwt_identity())

            # Trigger deal lastTouch update by logging a deal audit entry
            cursor.execute('SELECT deal_id, subject FROM activities WHERE id = %s', (activity_id,))
            d_row = cursor.fetchone()
            deal_id = d_row[0] if d_row else None
            task_name = d_row[1] if d_row else 'task'

            if deal_id:
                log_audit(conn, 'deal', deal_id, f'task_status:{task_name}', old_status, new_status, get_jwt_identity())

        conn.commit()
        return jsonify({'message': 'Activity status updated', 'dealId': deal_id})
    finally:
        close_connection(conn)


@app.route('/api/activities/<task_id>/reassign', methods=['PATCH'])
@jwt_required()
def reassign_activity(task_id):
    data = request.get_json()
    new_owner_id = data.get('newOwnerId')
    if not new_owner_id:
        return jsonify({'error': 'newOwnerId required'}), 400

    claims = get_jwt()
    user_role = claims.get('role', '')
    if ROLE_RANK.get(user_role, 0) <= 40:
        return jsonify({'error': 'Only managers can reassign tasks'}), 403

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('''
            SELECT a.owner_id, l.branch
            FROM activities a
            LEFT JOIN deals d ON a.deal_id = d.id
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            WHERE a.id = %s
        ''', (task_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Task not found'}), 404

        old_owner_id, task_branch = row

        cursor.execute('SELECT name, role, branch FROM team WHERE id = %s', (new_owner_id,))
        target = cursor.fetchone()
        if not target:
            return jsonify({'error': 'Target user not found'}), 404

        # Target must rank strictly below the caller (mirrors reassign_lead)
        user_id = get_jwt_identity()
        if not can_assign(claims, user_id, new_owner_id, target[1]):
            return jsonify({'error': f'Cannot reassign to a user with equal or higher role ({target[1]})'}), 403

        target_branch = (target[2] or '').lower()
        if task_branch and target_branch and task_branch.lower() != target_branch:
            return jsonify({'error': 'Cannot reassign task to an SR from a different branch'}), 403

        # Full customer handover: move lead + company + all its deals + all their tasks
        cursor.execute(
            'SELECT company_id FROM deals WHERE id = (SELECT deal_id FROM activities WHERE id = %s)',
            (task_id,)
        )
        anchor_row = cursor.fetchone()
        anchor_id = anchor_row[0] if anchor_row else None

        if anchor_id:
            cursor.execute('UPDATE leads SET owner_id = %s, reassigned_at = CURRENT_DATE WHERE id = %s', (new_owner_id, anchor_id))
            cursor.execute('UPDATE companies SET owner_id = %s WHERE id = %s', (new_owner_id, anchor_id))
            cursor.execute('UPDATE deals SET owner_id = %s WHERE company_id = %s', (new_owner_id, anchor_id))
            cursor.execute('''
                UPDATE activities SET owner_id = %s
                WHERE deal_id IN (SELECT id FROM (SELECT id FROM deals WHERE company_id = %s) AS d)
            ''', (new_owner_id, anchor_id))
            log_audit(conn, 'lead', anchor_id, 'reassign_handover', str(old_owner_id), str(new_owner_id), get_jwt_identity())
        else:
            # Manual task with no linked deal — move just this task
            cursor.execute('UPDATE activities SET owner_id = %s WHERE id = %s', (new_owner_id, task_id))

        log_audit(conn, 'activity', task_id, 'reassign', str(old_owner_id), str(new_owner_id), get_jwt_identity())
        conn.commit()
        return jsonify({'success': True, 'newOwner': target[0]})
    finally:
        close_connection(conn)


# ─── Deal Audit Logs ─────────────────────────────────────────────────────────

@app.route('/api/deals/<deal_id>/audit', methods=['GET'])
@jwt_required()
def get_deal_audit_logs(deal_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT a.id, a.entity_type AS entityType, a.entity_id AS entityId, a.action,
                   a.old_value AS oldValue, a.new_value AS newValue, a.changed_at AS changedAt,
                   t.name AS changedBy
            FROM audit_log a
            LEFT JOIN team t ON a.user_id = t.id
            WHERE a.entity_type = 'deal' AND a.entity_id = %s
            ORDER BY a.changed_at DESC
        """, (deal_id,))
        logs = rows_to_list(cursor)
        for log in logs:
            log['changedAt'] = to_pht(log.get('changedAt'))
        return jsonify(logs)
    finally:
        close_connection(conn)


# ─── Dashboard KPIs ───────────────────────────────────────────────────────────

@app.route('/api/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()

        # Branch scope for joined-leads queries (l alias) and direct leads queries
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        direct_parts, _, direct_params = build_scope(claims, branch, col='LOWER(TRIM(branch))', requested_region=region)

        # Build WHERE fragments for joined (deals+leads) and direct (leads) queries
        deal_where = list(scope_parts) + ['l.branch IS NOT NULL']
        deal_params = list(scope_params)
        if restrict_owner:
            deal_where.append('d.owner_id = %s')
            deal_params.append(user_id)
        deal_where_sql = 'WHERE ' + ' AND '.join(deal_where)

        lead_where = ["DATE_FORMAT(created_at, '%Y-%m') = DATE_FORMAT(CURDATE(), '%Y-%m')"] + list(direct_parts)
        lead_params = list(direct_params)
        if restrict_owner:
            lead_where.append('owner_id = %s')
            lead_params.append(user_id)

        cursor.execute(
            "SELECT COUNT(*) FROM leads WHERE " + " AND ".join(lead_where),
            tuple(lead_params),
        )
        new_leads = cursor.fetchone()[0]

        cursor.execute(f'''
            SELECT COUNT(*) FROM deals d
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            {deal_where_sql} AND d.stage NOT IN ('Closed Won', 'Closed Lost')
        ''', tuple(deal_params))
        active_deals = cursor.fetchone()[0]

        cursor.execute(f'''
            SELECT d.stage, COUNT(*) AS count, SUM(d.value) AS value
            FROM deals d
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            {deal_where_sql}
            GROUP BY d.stage
        ''', tuple(deal_params))
        deals_per_stage = rows_to_list(cursor)

        total_leads_where = ['1=1'] + list(direct_parts)
        total_leads_params = list(direct_params)
        if restrict_owner:
            total_leads_where.append('owner_id = %s')
            total_leads_params.append(user_id)

        cursor.execute(
            'SELECT COUNT(*) FROM leads WHERE ' + ' AND '.join(total_leads_where),
            tuple(total_leads_params),
        )
        total_leads = cursor.fetchone()[0]

        cursor.execute(
            "SELECT COUNT(*) FROM leads WHERE status = 'Converted' AND " + ' AND '.join(total_leads_where),
            tuple(total_leads_params),
        )
        converted = cursor.fetchone()[0]
        conversion_rate = round((converted / total_leads * 100), 1) if total_leads else 0

        cursor.execute(f'''
            SELECT COALESCE(SUM(d.value), 0) FROM deals d
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            {deal_where_sql} AND d.stage NOT IN ('Closed Won', 'Closed Lost')
        ''', tuple(deal_params))
        pipeline_value = float(cursor.fetchone()[0])

        return jsonify({
            'newLeads':       new_leads,
            'activeDeals':    active_deals,
            'dealsPerStage':  deals_per_stage,
            'conversionRate': conversion_rate,
            'pipelineValue':  pipeline_value,
        })
    finally:
        close_connection(conn)



# ─── Team: Profile ────────────────────────────────────────────────────────────

@app.route('/api/team/profile/photo', methods=['POST'])
@jwt_required()
def upload_profile_photo():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_IMAGE_EXTENSIONS:
        return jsonify({'error': f'Image format not allowed. Allowed: {", ".join(sorted(ALLOWED_IMAGE_EXTENSIONS))}'}), 400

    user_id = get_jwt_identity()
    file_id = str(uuid.uuid4())
    stored_name = f"profile_{user_id}_{file_id}{ext}"
    file.save(os.path.join(PROFILE_PIC_FOLDER, stored_name))

    url = f'/api/uploads/profile-pics/{stored_name}'
    return jsonify({'message': 'File uploaded', 'url': url}), 200


@app.route('/api/uploads/profile-pics/<filename>', methods=['GET'])
def serve_profile_pic(filename):
    return send_from_directory(PROFILE_PIC_FOLDER, filename, as_attachment=False)


@app.route('/api/team/profile/photo', methods=['DELETE'])
@jwt_required()
def delete_profile_photo():
    user_id = get_jwt_identity()

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute("SELECT profile_pic FROM team WHERE id = %s", (user_id,))
        old_pic = cursor.fetchone()
        
        if old_pic and old_pic[0] and old_pic[0].startswith('/api/uploads/profile-pics/'):
            old_filename = old_pic[0].split('/')[-1]
            old_path = os.path.join(PROFILE_PIC_FOLDER, old_filename)
            if os.path.exists(old_path):
                try: os.remove(old_path)
                except Exception as e:
                    print(f"Failed to remove old profile pic: {e}")

        cursor.execute('UPDATE team SET profile_pic = NULL, profile_zoom = 1.0, profile_offset_y = 0.0, profile_offset_x = 0.0, profile_rotation = 0 WHERE id = %s', (user_id,))
        conn.commit()
        return jsonify({'message': 'Profile photo removed'}), 200
    finally:
        close_connection(conn)


@app.route('/api/team/profile/photo/adjust', methods=['PUT'])
@jwt_required()
def adjust_profile_photo():
    user_id     = get_jwt_identity()
    data        = request.get_json()
    new_pic_url = data.get('profilePic') 
    zoom        = data.get('zoom', 1.0)
    offset_y    = data.get('offsetY', 0.0)
    offset_x    = data.get('offsetX', 0.0)
    rotation    = data.get('rotation', 0)

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        if new_pic_url:
            # Confirming a new photo: Cleanup old one first
            cursor.execute("SELECT profile_pic FROM team WHERE id = %s", (user_id,))
            row = cursor.fetchone()
            if row and row[0] and row[0].startswith('/api/uploads/profile-pics/'):
                old_filename = row[0].split('/')[-1]
                old_path = os.path.join(PROFILE_PIC_FOLDER, old_filename)
                if row[0] != new_pic_url and os.path.exists(old_path):
                    try: os.remove(old_path)
                    except: pass
            
            cursor.execute(
                '''UPDATE team 
                   SET profile_pic = %s, profile_zoom = %s, profile_offset_y = %s, profile_offset_x = %s, profile_rotation = %s 
                   WHERE id = %s''', 
                (new_pic_url, zoom, offset_y, offset_x, rotation, user_id)
            )
        else:
            # Just adjusting current photo
            cursor.execute(
                'UPDATE team SET profile_zoom = %s, profile_offset_y = %s, profile_offset_x = %s, profile_rotation = %s WHERE id = %s', 
                (zoom, offset_y, offset_x, rotation, user_id)
            )

        conn.commit()
        return jsonify({'message': 'Profile updated successfully'}), 200
    finally:
        close_connection(conn)


@app.route('/api/team/profile/password', methods=['PUT'])
@jwt_required()
def update_user_password():
    user_id          = get_jwt_identity()
    data             = request.get_json()
    current_password = data.get('currentPassword', '').strip()
    new_password     = data.get('newPassword', '').strip()

    if not current_password or not new_password:
        return jsonify({'error': 'Both current and new passwords are required.'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        # Fetch hashed password for verification
        cursor.execute("SELECT password FROM team WHERE id = %s", (user_id,))
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'User not found.'}), 404

        valid_password, should_rehash = verify_password(row[0], current_password)
        if not valid_password:
            return jsonify({'error': 'Current password is incorrect.'}), 400

        # Always hash the new password before storing
        hashed_new = generate_password_hash(new_password)
        cursor.execute('UPDATE team SET password = %s WHERE id = %s', (hashed_new, user_id))
        conn.commit()

        return jsonify({'message': 'Password updated successfully.'})
    finally:
        close_connection(conn)


# ─── Admin: Profile ───────────────────────────────────────────────────────────

@app.route('/api/admin/profile', methods=['PUT'])
@jwt_required()
@admin_required
def update_admin_profile():
    data             = request.get_json()
    admin_id         = data.get('id')
    current_password = data.get('currentPassword', '').strip()
    new_username     = data.get('newUsername', '').strip()
    new_password     = data.get('newPassword', '').strip()

    if not admin_id or not current_password:
        return jsonify({'error': 'Current password is required.'}), 400
    if not new_username and not new_password:
        return jsonify({'error': 'Provide a new username or new password.'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        # Fetch hashed password for verification — never compare plaintext
        cursor.execute(
            "SELECT id, password FROM team WHERE id = %s AND role = 'Admin'",
            (admin_id,)
        )
        row = cursor.fetchone()
        if not row:
            return jsonify({'error': 'Current password is incorrect.'}), 401

        valid_password, should_rehash = verify_password(row[1], current_password)
        if not valid_password:
            return jsonify({'error': 'Current password is incorrect.'}), 401

        if should_rehash:
            cursor.execute('UPDATE team SET password = %s WHERE id = %s', (generate_password_hash(current_password), admin_id))
            conn.commit()

        if new_username:
            cursor.execute('SELECT id FROM team WHERE username = %s AND id != %s', (new_username, admin_id))
            if cursor.fetchone():
                return jsonify({'error': 'Username is already taken.'}), 409

        # Always hash the new password before storing
        hashed_new = generate_password_hash(new_password) if new_password else None

        if new_username and hashed_new:
            cursor.execute('UPDATE team SET username = %s, password = %s WHERE id = %s', (new_username, hashed_new, admin_id))
        elif new_username:
            cursor.execute('UPDATE team SET username = %s WHERE id = %s', (new_username, admin_id))
        else:
            cursor.execute('UPDATE team SET password = %s WHERE id = %s', (hashed_new, admin_id))

        conn.commit()

        cursor.execute('SELECT id, username, name, email, role, branch FROM team WHERE id = %s', (admin_id,))
        row = cursor.fetchone()
        user = {'id': row[0], 'username': row[1], 'name': row[2], 'email': row[3], 'role': row[4], 'branch': row[5]}
        return jsonify({'message': 'Profile updated successfully.', 'user': user}), 200
    finally:
        close_connection(conn)


# ─── Admin: Analytics ─────────────────────────────────────────────────────────

@app.route('/api/admin/analytics', methods=['GET'])
@jwt_required()
@admin_required
def admin_analytics():
    branch_filter  = request.args.get('branch', '').strip()
    region_filter  = request.args.get('region', '').strip()
    region_branches = REGION_BRANCHES.get(region_filter, []) if (not branch_filter and region_filter) else []
    audit_limit   = min(int(request.args.get('auditLimit',  20)), 100)
    audit_offset  = int(request.args.get('auditOffset', 0))
    audit_entity  = request.args.get('auditEntity', '').strip()
    audit_from    = request.args.get('auditFrom',   '').strip()

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        # Users per branch (exclude Headquarters — admin-only, not a sales branch)
        if branch_filter:
            cursor.execute("SELECT branch, COUNT(*) AS count FROM team WHERE branch != 'Headquarters' AND branch = %s GROUP BY branch ORDER BY branch", (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f"SELECT branch, COUNT(*) AS count FROM team WHERE branch != 'Headquarters' AND branch IN ({in_ph}) GROUP BY branch ORDER BY branch", region_branches)
        else:
            cursor.execute("SELECT branch, COUNT(*) AS count FROM team WHERE branch != 'Headquarters' GROUP BY branch ORDER BY branch")
        users_per_branch = rows_to_list(cursor)

        # Role distribution
        if branch_filter:
            cursor.execute("SELECT role, COUNT(*) AS count FROM team WHERE branch != 'Headquarters' AND branch = %s GROUP BY role ORDER BY role", (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f"SELECT role, COUNT(*) AS count FROM team WHERE branch != 'Headquarters' AND branch IN ({in_ph}) GROUP BY role ORDER BY role", region_branches)
        else:
            cursor.execute("SELECT role, COUNT(*) AS count FROM team WHERE branch != 'Headquarters' GROUP BY role ORDER BY role")
        role_distribution = rows_to_list(cursor)

        # Leads per branch
        if branch_filter:
            cursor.execute('SELECT branch, COUNT(*) AS total, SUM(status = "Converted") AS converted FROM leads WHERE branch = %s GROUP BY branch ORDER BY branch', (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f'SELECT branch, COUNT(*) AS total, SUM(status = "Converted") AS converted FROM leads WHERE branch IN ({in_ph}) GROUP BY branch ORDER BY branch', region_branches)
        else:
            cursor.execute('SELECT branch, COUNT(*) AS total, SUM(status = "Converted") AS converted FROM leads GROUP BY branch ORDER BY branch')
        leads_per_branch = rows_to_list(cursor)

        # Lead status distribution
        if branch_filter:
            cursor.execute(
                'SELECT status, COUNT(*) AS count FROM leads WHERE branch = %s GROUP BY status ORDER BY count DESC',
                (branch_filter,)
            )
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(
                f'SELECT status, COUNT(*) AS count FROM leads WHERE branch IN ({in_ph}) GROUP BY status ORDER BY count DESC',
                region_branches
            )
        else:
            cursor.execute(
                'SELECT status, COUNT(*) AS count FROM leads GROUP BY status ORDER BY count DESC'
            )
        lead_status_dist = rows_to_list(cursor)

        # Deals per branch — active deal_count + pipeline_value, plus win_rate and avg_deal_value across all deals
        deals_sql = '''
            SELECT l.branch,
                   COUNT(CASE WHEN d.stage NOT IN ('Closed Won','Closed Lost') THEN 1 END) AS deal_count,
                   COALESCE(SUM(CASE WHEN d.stage NOT IN ('Closed Won','Closed Lost') THEN d.value ELSE 0 END), 0) AS pipeline_value,
                   ROUND(
                       SUM(d.stage = 'Closed Won') /
                       NULLIF(SUM(d.stage IN ('Closed Won','Closed Lost')), 0) * 100, 1
                   ) AS win_rate,
                   ROUND(AVG(d.value), 0) AS avg_deal_value
            FROM deals d
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            {where}
            GROUP BY l.branch
            ORDER BY l.branch
        '''
        if branch_filter:
            cursor.execute(deals_sql.format(where='WHERE l.branch = %s'), (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(deals_sql.format(where=f'WHERE l.branch IN ({in_ph})'), region_branches)
        else:
            cursor.execute(deals_sql.format(where=''))
        deals_per_branch = rows_to_list(cursor)

        # Top SRs — leads, converted, deals won
        # Uses owner_name as fallback for Excel-imported leads where owner_id is NULL
        if branch_filter:
            cursor.execute('''
                SELECT COALESCE(t.name, l.owner_name)   AS name,
                       COALESCE(t.branch, l.branch)     AS branch,
                       COUNT(DISTINCT l.id)                                            AS leads_count,
                       COUNT(DISTINCT CASE WHEN l.status = 'Converted' THEN l.id END) AS converted,
                       COUNT(DISTINCT CASE WHEN d.stage = 'Closed Won' THEN d.id END) AS deals_won
                FROM leads l
                LEFT JOIN team t ON l.owner_id = t.id
                LEFT JOIN deals d ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE (l.owner_id IS NOT NULL OR (l.owner_name IS NOT NULL AND TRIM(l.owner_name) != ''))
                  AND (t.id IS NULL OR t.role IN ('Sales Representative', 'Branch Account', 'Sales Rep'))
                  AND LOWER(TRIM(l.branch)) = LOWER(TRIM(%s))
                GROUP BY COALESCE(t.name, l.owner_name), COALESCE(t.branch, l.branch)
                ORDER BY leads_count DESC
                LIMIT 20
            ''', (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f'''
                SELECT COALESCE(t.name, l.owner_name)   AS name,
                       COALESCE(t.branch, l.branch)     AS branch,
                       COUNT(DISTINCT l.id)                                            AS leads_count,
                       COUNT(DISTINCT CASE WHEN l.status = 'Converted' THEN l.id END) AS converted,
                       COUNT(DISTINCT CASE WHEN d.stage = 'Closed Won' THEN d.id END) AS deals_won
                FROM leads l
                LEFT JOIN team t ON l.owner_id = t.id
                LEFT JOIN deals d ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE (l.owner_id IS NOT NULL OR (l.owner_name IS NOT NULL AND TRIM(l.owner_name) != ''))
                  AND (t.id IS NULL OR t.role IN ('Sales Representative', 'Branch Account', 'Sales Rep'))
                  AND LOWER(TRIM(l.branch)) IN ({in_ph})
                GROUP BY COALESCE(t.name, l.owner_name), COALESCE(t.branch, l.branch)
                ORDER BY leads_count DESC
                LIMIT 20
            ''', region_branches)
        else:
            cursor.execute('''
                SELECT COALESCE(t.name, l.owner_name)   AS name,
                       COALESCE(t.branch, l.branch)     AS branch,
                       COUNT(DISTINCT l.id)                                            AS leads_count,
                       COUNT(DISTINCT CASE WHEN l.status = 'Converted' THEN l.id END) AS converted,
                       COUNT(DISTINCT CASE WHEN d.stage = 'Closed Won' THEN d.id END) AS deals_won
                FROM leads l
                LEFT JOIN team t ON l.owner_id = t.id
                LEFT JOIN deals d ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE (l.owner_id IS NOT NULL OR (l.owner_name IS NOT NULL AND TRIM(l.owner_name) != ''))
                  AND (t.id IS NULL OR t.role IN ('Sales Representative', 'Branch Account', 'Sales Rep'))
                GROUP BY COALESCE(t.name, l.owner_name), COALESCE(t.branch, l.branch)
                ORDER BY leads_count DESC
                LIMIT 20
            ''')
        top_srs = rows_to_list(cursor)

        # Totals
        if branch_filter:
            cursor.execute("SELECT COUNT(*) FROM team WHERE branch != 'Headquarters' AND branch = %s", (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f"SELECT COUNT(*) FROM team WHERE branch != 'Headquarters' AND branch IN ({in_ph})", region_branches)
        else:
            cursor.execute("SELECT COUNT(*) FROM team WHERE branch != 'Headquarters'")
        total_users = cursor.fetchone()[0]

        if branch_filter:
            cursor.execute("SELECT COUNT(*) FROM leads WHERE branch = %s", (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f"SELECT COUNT(*) FROM leads WHERE branch IN ({in_ph})", region_branches)
        else:
            cursor.execute("SELECT COUNT(*) FROM leads")
        total_leads = cursor.fetchone()[0]

        if branch_filter:
            cursor.execute('''
                SELECT COUNT(*)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage NOT IN ('Closed Won', 'Closed Lost') AND l.branch = %s
            ''', (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f'''
                SELECT COUNT(*)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage NOT IN ('Closed Won', 'Closed Lost') AND l.branch IN ({in_ph})
            ''', region_branches)
        else:
            cursor.execute('''
                SELECT COUNT(*)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage NOT IN ('Closed Won', 'Closed Lost')
            ''')
        active_deals = cursor.fetchone()[0]

        if branch_filter:
            cursor.execute('''
                SELECT COALESCE(SUM(d.value),0)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage NOT IN ('Closed Won', 'Closed Lost') AND l.branch = %s
            ''', (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f'''
                SELECT COALESCE(SUM(d.value),0)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage NOT IN ('Closed Won', 'Closed Lost') AND l.branch IN ({in_ph})
            ''', region_branches)
        else:
            cursor.execute('''
                SELECT COALESCE(SUM(d.value),0)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage NOT IN ('Closed Won', 'Closed Lost')
            ''')
        pipeline_value = float(cursor.fetchone()[0])

        if branch_filter:
            cursor.execute('''
                SELECT COUNT(*)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage = 'Closed Won' AND l.branch = %s
            ''', (branch_filter,))
        elif region_branches:
            in_ph = ', '.join(['%s'] * len(region_branches))
            cursor.execute(f'''
                SELECT COUNT(*)
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                WHERE d.stage = 'Closed Won' AND l.branch IN ({in_ph})
            ''', region_branches)
        else:
            cursor.execute("SELECT COUNT(*) FROM deals WHERE stage = 'Closed Won'")
        closed_won = cursor.fetchone()[0]

        # Audit log — paginated + filterable
        audit_filters = []
        audit_params  = []
        if audit_entity:
            audit_filters.append('entity_type = %s')
            audit_params.append(audit_entity)
        if audit_from:
            audit_filters.append('changed_at >= %s')
            audit_params.append(audit_from)
        audit_where = ('WHERE ' + ' AND '.join(audit_filters)) if audit_filters else ''

        cursor.execute(f'SELECT COUNT(*) FROM audit_log {audit_where}', audit_params)
        audit_total = cursor.fetchone()[0]

        cursor.execute(
            f'SELECT * FROM audit_log {audit_where} ORDER BY changed_at DESC LIMIT %s OFFSET %s',
            audit_params + [audit_limit, audit_offset]
        )
        audit_log = rows_to_list(cursor)

        return jsonify({
            'usersPerBranch':   users_per_branch,
            'roleDistribution': role_distribution,
            'leadsPerBranch':   leads_per_branch,
            'dealsPerBranch':   deals_per_branch,
            'topSRs':           top_srs,
            'leadStatusDist':   lead_status_dist,
            'totals': {
                'users':         total_users,
                'leads':         total_leads,
                'activeDeals':   active_deals,
                'pipelineValue': pipeline_value,
                'closedWon':     closed_won,
            },
            'auditLog':   audit_log,
            'auditTotal': audit_total,
        })
    finally:
        close_connection(conn)


@app.route('/api/admin/audit-log', methods=['GET'])
@jwt_required()
@admin_required
def admin_audit_log():
    limit      = min(int(request.args.get('limit',  20)), 100)
    offset     = int(request.args.get('offset', 0))
    entity     = request.args.get('entity', '').strip()
    date_from  = request.args.get('from',   '').strip()
    date_to    = request.args.get('to',     '').strip()
    user_id    = request.args.get('userId', '').strip()

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        filters = []
        params  = []
        if entity:
            filters.append('entity_type = %s');   params.append(entity)
        if date_from:
            filters.append('changed_at >= %s');   params.append(date_from)
        if date_to:
            filters.append('changed_at <= %s');   params.append(date_to)
        if user_id:
            filters.append('user_id = %s');       params.append(user_id)

        where = ('WHERE ' + ' AND '.join(filters)) if filters else ''

        cursor.execute(f'SELECT COUNT(*) FROM audit_log {where}', params)
        total = cursor.fetchone()[0]

        cursor.execute(
            f'SELECT * FROM audit_log {where} ORDER BY changed_at DESC LIMIT %s OFFSET %s',
            params + [limit, offset]
        )
        logs = rows_to_list(cursor)
        for log in logs:
            log['changed_at'] = to_pht(log.get('changed_at'))

        return jsonify({'logs': logs, 'total': total})
    finally:
        close_connection(conn)


# ─── Admin: Excel Export ──────────────────────────────────────────────────────

@app.route('/api/admin/export/<report>', methods=['GET'])
@jwt_required()
@admin_required
def admin_export(report):
    if report not in ('branch-overview', 'audit-log'):
        return jsonify({'error': 'Unknown report'}), 404

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        wb = Workbook()
        ws = wb.active
        
        # TDT Theme Styles
        header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        center_align = Alignment(horizontal='center')

        if report == 'branch-overview':
            ws.title = "Branch Overview"
            headers = ['Branch', 'Active Deals', 'Pipeline Value', 'Win Rate (%)', 'Avg Deal Value']
            ws.append(headers)
            
            # Pre-populate with all official branches to ensure zero-data branches appear
            branch_data = {}
            for branches in REGION_BRANCHES.values():
                for bname in branches:
                    branch_data[bname] = [bname, 0, 0, None, None]

            cursor.execute('''
                SELECT l.branch,
                       COUNT(CASE WHEN d.stage NOT IN ('Closed Won','Closed Lost') THEN 1 END) AS deal_count,
                       COALESCE(SUM(CASE WHEN d.stage NOT IN ('Closed Won','Closed Lost') THEN d.value ELSE 0 END), 0) AS pipeline_value,
                       ROUND(
                           SUM(d.stage = 'Closed Won') /
                           NULLIF(SUM(d.stage IN ('Closed Won','Closed Lost')), 0) * 100, 1
                       ) AS win_rate,
                       ROUND(AVG(d.value), 0) AS avg_deal_value
                FROM deals d
                LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
                GROUP BY l.branch
            ''')
            rows = cursor.fetchall()
            for row in rows:
                if row[0] in branch_data:
                    branch_data[row[0]] = list(row)
                
            # Append sorted rows
            for bname in sorted(branch_data.keys()):
                ws.append(branch_data[bname])
                
            # Formatting
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                
            # Number Formatting (Thousand Separators)
            for row in ws.iter_rows(min_row=2, min_col=2, max_col=5):
                if row[0].value is not None: row[0].number_format = '#,##0'       # Active Deals
                if row[1].value is not None: row[1].number_format = '#,##0'       # Pipeline Value
                if row[3].value is not None: row[3].number_format = '#,##0'       # Avg Deal Value
                
            # Column widths
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 20
            
            filename = 'branch-overview.xlsx'

        else:  # audit-log
            ws.title = "Audit Log"
            headers = ['ID', 'Entity Type', 'Entity ID', 'Action', 'Old Value', 'New Value', 'Changed At', 'User ID']
            ws.append(headers)

            entity    = request.args.get('entity', '').strip()
            date_from = request.args.get('from',   '').strip()
            date_to   = request.args.get('to',     '').strip()
            user_id   = request.args.get('userId', '').strip()

            filters = []
            params  = []
            if entity:
                filters.append('entity_type = %s');  params.append(entity)
            if date_from:
                filters.append('changed_at >= %s');  params.append(date_from)
            if date_to:
                filters.append('changed_at <= %s');  params.append(date_to)
            if user_id:
                filters.append('user_id = %s');      params.append(user_id)

            where = ('WHERE ' + ' AND '.join(filters)) if filters else ''
            cursor.execute(
                f'SELECT * FROM audit_log {where} ORDER BY changed_at DESC LIMIT 5000',
                params
            )
            rows = cursor.fetchall()
            for row in rows:
                # Convert datetime to string for Excel compatibility if needed
                lrow = list(row)
                if isinstance(lrow[6], datetime):
                    lrow[6] = lrow[6].strftime('%Y-%m-%d %H:%M:%S')
                ws.append(lrow)
                
            # Formatting
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                
            # Column widths
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws.column_dimensions[col].width = 20
                
            filename = 'audit-log.xlsx'

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    finally:
        close_connection(conn)


# ─── Admin: User Management ───────────────────────────────────────────────────

@app.route('/api/admin/users', methods=['GET'])
@jwt_required()
@admin_required
def admin_get_users():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch')
        if branch and branch != 'Headquarters':
            cursor.execute(
                "SELECT id, username, name, email, role, branch FROM team WHERE branch = %s AND branch != 'Headquarters' ORDER BY branch, name",
                (branch,)
            )
        else:
            cursor.execute(
                "SELECT id, username, name, email, role, branch FROM team WHERE branch != 'Headquarters' ORDER BY branch, name"
            )
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/admin/users', methods=['POST'])
@jwt_required()
@admin_required
def admin_create_user():
    data = request.get_json()
    required = ['username', 'password', 'name', 'branch']
    if not all(data.get(k, '').strip() for k in required):
        return jsonify({'error': 'username, password, name, and branch are required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        # Hash the password before storing — never store plaintext
        hashed_password = generate_password_hash(data['password'])
        cursor.execute(
            'INSERT INTO team (username, password, name, email, role, branch) VALUES (%s, %s, %s, %s, %s, %s)',
            (
                data['username'].strip(),
                hashed_password,
                data['name'].strip(),
                data.get('email', '').strip(),
                data.get('role', 'Branch Account'),
                data['branch'],
            )
        )
        conn.commit()
        return jsonify({'message': 'User created', 'id': cursor.lastrowid}), 201
    except Exception as e:
        if 'Duplicate entry' in str(e):
            return jsonify({'error': 'Username already exists'}), 409
        # Log actual error securely, return generic message
        print(f"admin_create_user error: {traceback.format_exc()}")
        return jsonify({'error': 'Failed to create user'}), 500
    finally:
        close_connection(conn)


@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@jwt_required()
@admin_required
def admin_update_user(user_id):
    data = request.get_json()
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM team WHERE id = %s', (user_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'User not found'}), 404

        fields, values = [], []
        # Only allow known safe columns — never pass col names from user input
        for col in ['name', 'email', 'role', 'branch', 'username']:
            if col in data:
                fields.append(f'{col} = %s')
                values.append(data[col])
        if data.get('password'):
            # Hash updated password before storing
            fields.append('password = %s')
            values.append(generate_password_hash(data['password']))

        if not fields:
            return jsonify({'error': 'No fields to update'}), 400

        values.append(user_id)
        cursor.execute(f'UPDATE team SET {", ".join(fields)} WHERE id = %s', values)
        conn.commit()
        return jsonify({'message': 'User updated'})
    finally:
        close_connection(conn)


@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def admin_delete_user(user_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT id FROM team WHERE id = %s', (user_id,))
        if not cursor.fetchone():
            return jsonify({'error': 'User not found'}), 404
        cursor.execute('DELETE FROM team WHERE id = %s', (user_id,))
        conn.commit()
        return jsonify({'message': 'User deleted'})
    finally:
        close_connection(conn)


# ─── Excel Customer Import ────────────────────────────────────────────────────

@app.route('/api/admin/import/customers', methods=['POST'])
@jwt_required()
@admin_required
def import_customers():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith(('.xlsx', '.xls')):
        return jsonify({'error': 'Only .xlsx and .xls files are accepted'}), 400

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb['Sheet1']
    except Exception as e:
        return jsonify({'error': f'Could not read file: {str(e)}'}), 400

    branch_to_region = {}
    for region, branches in REGION_BRANCHES.items():
        for b in branches:
            branch_to_region[b] = region

    inserted = 0
    skipped  = 0
    errors   = []

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Right-side clean columns (0-indexed): 22=Customer, 23=Bill to, 25=Main Phone, 26=Rep, 27=Class
            if not row or len(row) < 28:
                skipped += 1
                continue

            customer_name = str(row[22]).strip().replace('\xa0', '').strip() if row[22] else ''
            address       = str(row[23]).strip() if row[23] else ''
            contact_num   = _parse_phone(row[25])
            owner_name    = str(row[26]).strip() if row[26] else ''
            branch        = str(row[27]).strip() if row[27] else ''

            if not customer_name or not branch:
                if customer_name or branch:
                    errors.append(f'Row {i}: missing {"branch" if not branch else "customer name"}')
                skipped += 1
                continue

            region = branch_to_region.get(branch, '')

            try:
                cursor.execute('SAVEPOINT import_row')

                cursor.execute(
                    'SELECT id FROM leads WHERE customer_name = %s AND branch = %s LIMIT 1',
                    (customer_name, branch)
                )
                existing_lead = cursor.fetchone()
                if existing_lead:
                    existing_lead_id = existing_lead[0]
                    contact_person = _parse_contact_name(row[24])
                    cursor.execute(
                        'SELECT id, name FROM contacts WHERE company_id = %s ORDER BY created_at LIMIT 1',
                        (existing_lead_id,)
                    )
                    existing_contact = cursor.fetchone()
                    if not existing_contact:
                        if contact_person or contact_num:
                            cursor.execute('''
                                INSERT INTO contacts (id, name, company_id, phone, role, status)
                                VALUES (%s, %s, %s, %s, 'Primary Contact', 'Active')
                            ''', (str(uuid.uuid4()), contact_person or '—', existing_lead_id, contact_num))
                    elif existing_contact[1] in ('', customer_name):
                        # Auto-created with company-name fallback or blank — correct it
                        cursor.execute(
                            'UPDATE contacts SET name=%s, phone=%s WHERE id=%s',
                            (contact_person or '—', contact_num, existing_contact[0])
                        )
                    cursor.execute('RELEASE SAVEPOINT import_row')
                    skipped += 1
                    continue

                lead_id = str(uuid.uuid4())

                cursor.execute('''
                    INSERT INTO leads (id, customer_name, contact_num, address, region, branch, owner_name, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'New')
                ''', (lead_id, customer_name, contact_num, address, region, branch, owner_name))

                # companies table only stores id + name (address/branch/region live on leads)
                cursor.execute('''
                    INSERT INTO companies (id, name)
                    VALUES (%s, %s)
                    ON DUPLICATE KEY UPDATE name=name
                ''', (lead_id, customer_name))

                # column 24 = Primary Contact person name; col 25 = phone
                contact_person = _parse_contact_name(row[24])
                if contact_person or contact_num:
                    cursor.execute('''
                        INSERT INTO contacts (id, name, company_id, phone, role, status)
                        VALUES (%s, %s, %s, %s, 'Primary Contact', 'Active')
                    ''', (str(uuid.uuid4()), contact_person or '—', lead_id, contact_num))

                cursor.execute('RELEASE SAVEPOINT import_row')
                inserted += 1

            except Exception as row_err:
                cursor.execute('ROLLBACK TO SAVEPOINT import_row')
                errors.append(f'Row {i} ({customer_name[:40]}): {str(row_err)[:100]}')
                skipped += 1

            if inserted % 500 == 0:
                conn.commit()

        conn.commit()
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)

    return jsonify({'inserted': inserted, 'skipped': skipped, 'errors': errors[:50]})


# ─── Deal Attachments ─────────────────────────────────────────────────────────

@app.route('/api/deals/<deal_id>/attachments', methods=['POST'])
@jwt_required()
def upload_deal_attachment(deal_id):
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if file.content_type not in ('application/pdf', 'application/octet-stream') and not file.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Only PDF files are allowed'}), 400

    safe_name = secure_filename(file.filename)
    file_id = str(uuid.uuid4())
    stored_name = f"{file_id}_{safe_name}"
    file.save(os.path.join(UPLOAD_FOLDER, stored_name))

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO deal_attachments (id, deal_id, filename, label) VALUES (%s, %s, %s, %s)',
            (file_id, deal_id, stored_name, safe_name),
        )
        conn.commit()
        return jsonify({'id': file_id, 'filename': stored_name, 'label': safe_name}), 201
    finally:
        close_connection(conn)


@app.route('/api/deals/<deal_id>/attachments', methods=['GET'])
@jwt_required()
def get_deal_attachments(deal_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            'SELECT id, filename, label, uploaded_at FROM deal_attachments WHERE deal_id = %s ORDER BY uploaded_at DESC',
            (deal_id,),
        )
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/uploads/<filename>', methods=['GET'])
@jwt_required()
def serve_attachment(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


@app.after_request
def set_security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    return response


@app.route('/api/deal-contacts', methods=['GET'])
@jwt_required()
def get_all_deal_contacts():
    """Return all deal_contacts with full contact info - single efficient query"""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        branch = request.args.get('branch', '')
        region = request.args.get('region', '')
        claims = get_jwt()
        user_id = get_jwt_identity()
        
        scope_parts, restrict_owner, scope_params = build_scope(claims, branch, requested_region=region)
        where_parts = list(scope_parts)
        params = list(scope_params)
        
        if restrict_owner:
            where_parts.append('d.owner_id = %s')
            params.append(user_id)
            
        where_sql = 'WHERE ' + ' AND '.join(where_parts) if where_parts else ''
        
        cursor.execute(f"""
            SELECT dc.deal_id, c.id, c.name, c.role, c.email, c.phone, dc.role AS deal_role
            FROM deal_contacts dc
            JOIN contacts c ON c.id = dc.contact_id
            JOIN deals d ON d.id = dc.deal_id
            LEFT JOIN leads l ON (d.lead_id = l.id OR d.company_id = l.id)
            {where_sql}
        """, tuple(params))
        return jsonify(rows_to_list(cursor))
    except Exception as e:
        print(f"Error in get_all_deal_contacts: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


@app.route('/api/deals/<deal_id>/contacts', methods=['GET'])
@jwt_required()
def get_deal_contacts(deal_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            '''
            SELECT c.id, c.name, c.role, c.email, c.phone, dc.role AS deal_role
            FROM contacts c
            INNER JOIN deal_contacts dc ON c.id = dc.contact_id
            WHERE dc.deal_id = %s
            ''',
            (deal_id,),
        )
        return jsonify(rows_to_list(cursor))
    except Exception as e:
        print(f"Error in get_deal_contacts: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


@app.route('/api/deals/<deal_id>/contacts', methods=['POST'])
@jwt_required()
def add_deal_contact(deal_id):
    data = request.get_json()
    contact_id = data.get('contactId')
    role = data.get('role', 'Primary')
    if not contact_id:
        return jsonify({'error': 'contactId is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        
        # Get old role if exists for audit
        cursor.execute("SELECT role FROM deal_contacts WHERE deal_id = %s AND contact_id = %s", (deal_id, contact_id))
        old_row = cursor.fetchone()
        old_role = old_row[0] if old_row else None

        cursor.execute(
            '''INSERT INTO deal_contacts (deal_id, contact_id, role) 
               VALUES (%s, %s, %s)
               ON DUPLICATE KEY UPDATE role = VALUES(role)''',
            (deal_id, contact_id, role),
        )

        # Log audit
        cursor.execute("SELECT name FROM contacts WHERE id = %s", (contact_id,))
        c_row = cursor.fetchone()
        c_name = c_row[0] if c_row else "Unknown Contact"
        
        user_id = get_jwt_identity()
        if old_role:
            if old_role != role:
                log_audit(conn, 'deal', deal_id, f'contact_role_change:{c_name}', old_role, role, user_id)
        else:
            log_audit(conn, 'deal', deal_id, 'contact_added', None, f"{c_name} ({role})", user_id)

        conn.commit()
        return jsonify({'message': 'Contact added to deal'}), 201
    except Exception as e:
        print(f"Error in add_deal_contact: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


@app.route('/api/deals/<deal_id>/contacts', methods=['DELETE'])
@jwt_required()
def remove_deal_contact(deal_id):
    data = request.get_json()
    contact_id = data.get('contactId')
    if not contact_id:
        return jsonify({'error': 'contactId is required'}), 400
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()

        # Get contact name for audit
        cursor.execute("SELECT name FROM contacts WHERE id = %s", (contact_id,))
        c_row = cursor.fetchone()
        c_name = c_row[0] if c_row else "Unknown Contact"

        cursor.execute(
            'DELETE FROM deal_contacts WHERE deal_id = %s AND contact_id = %s',
            (deal_id, contact_id),
        )
        
        log_audit(conn, 'deal', deal_id, 'contact_removed', c_name, None, get_jwt_identity())
        
        conn.commit()
        return jsonify({'message': 'Contact removed from deal'}), 200
    except Exception as e:
        print(f"Error in remove_deal_contact: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        close_connection(conn)


@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'healthy'}), 200


# ─── Celebration Music ─────────────────────────────────────────────────────────

@app.route('/api/admin/settings/celebration-music', methods=['GET'])
@jwt_required()
@admin_required
def admin_get_celebration_music():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT id, outcome, source_type, url, original_filename, is_active FROM celebration_music WHERE outcome IN (%s, %s) ORDER BY outcome', ('won', 'lost'))
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/admin/settings/celebration-music/<outcome>', methods=['POST'])
@jwt_required()
@admin_required
def admin_add_celebration_music(outcome):
    if outcome not in ('won', 'lost'):
        return jsonify({'error': 'Outcome must be "won" or "lost"'}), 400
    data = request.get_json()
    if not data or not data.get('url'):
        return jsonify({'error': 'url is required'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO celebration_music (outcome, url, source_type) VALUES (%s, %s, %s)',
            (outcome, data['url'], 'url'),
        )
        conn.commit()
        return jsonify({'id': cursor.lastrowid, 'message': f'{outcome} music added'}), 201
    finally:
        close_connection(conn)


@app.route('/api/admin/settings/celebration-music/upload', methods=['POST'])
@jwt_required()
@admin_required
def admin_upload_celebration_music():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    outcome = request.form.get('outcome')
    if outcome not in ('won', 'lost'):
        return jsonify({'error': 'outcome must be "won" or "lost"'}), 400

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ALLOWED_AUDIO_EXTENSIONS:
        return jsonify({'error': f'Audio format not allowed. Allowed: {", ".join(sorted(ALLOWED_AUDIO_EXTENSIONS))}'}), 400

    safe_name = secure_filename(file.filename)
    file_id = str(uuid.uuid4())
    stored_name = f"{file_id}{ext}"
    file.save(os.path.join(MUSIC_UPLOAD_FOLDER, stored_name))

    url = f'/api/uploads/celebration-music/{stored_name}'

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            'INSERT INTO celebration_music (outcome, url, source_type, original_filename, stored_filename) VALUES (%s, %s, %s, %s, %s)',
            (outcome, url, 'internal', safe_name, stored_name),
        )
        conn.commit()
        entry_id = cursor.lastrowid
    finally:
        close_connection(conn)

    return jsonify({
        'id': entry_id,
        'url': url,
        'originalFilename': safe_name,
        'storedFilename': stored_name,
    }), 201


@app.route('/api/admin/settings/celebration-music/entry/<int:entry_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def admin_delete_celebration_music_entry(entry_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('DELETE FROM celebration_music WHERE id = %s', (entry_id,))
        conn.commit()
        if cursor.rowcount == 0:
            return jsonify({'error': 'Entry not found'}), 404
        return jsonify({'message': 'Music entry deleted'})
    finally:
        close_connection(conn)


@app.route('/api/celebration-music', methods=['GET'])
@jwt_required()
def get_celebration_music():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute('SELECT outcome, source_type, url, original_filename FROM celebration_music WHERE is_active = 1')
        return jsonify(rows_to_list(cursor))
    finally:
        close_connection(conn)


@app.route('/api/uploads/celebration-music/<filename>', methods=['GET'])
def serve_celebration_music(filename):
    return send_from_directory(MUSIC_UPLOAD_FOLDER, filename, as_attachment=False)


# ─── Celebration Animation Settings ─────────────────────────────────────────

_VALID_ANIMATIONS = {'confetti', 'jojo', 'none'}

@app.route('/api/admin/settings/celebration-animation', methods=['GET'])
@jwt_required()
@admin_required
def admin_get_celebration_animation():
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT setting_key, setting_value FROM app_settings WHERE setting_key IN (%s, %s)",
            ('celebration_animation_won', 'celebration_animation_lost')
        )
        rows = {r['setting_key']: r['setting_value'] for r in rows_to_list(cursor)}
        return jsonify({
            'won':  rows.get('celebration_animation_won',  'confetti'),
            'lost': rows.get('celebration_animation_lost', 'confetti'),
        })
    finally:
        close_connection(conn)


@app.route('/api/admin/settings/celebration-animation', methods=['PUT'])
@jwt_required()
@admin_required
def admin_put_celebration_animation():
    data = request.get_json(silent=True) or {}
    updates = {}
    for outcome in ('won', 'lost'):
        val = data.get(outcome)
        if val is not None:
            if val not in _VALID_ANIMATIONS:
                return jsonify({'error': f'Invalid animation value: {val}'}), 400
            updates[f'celebration_animation_{outcome}'] = val
    if not updates:
        return jsonify({'error': 'No valid fields to update'}), 400

    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        for key, value in updates.items():
            cursor.execute(
                "INSERT INTO app_settings (setting_key, setting_value) VALUES (%s, %s) "
                "ON DUPLICATE KEY UPDATE setting_value = %s",
                (key, value, value)
            )
        conn.commit()
        return jsonify({'message': 'Animation settings updated'})
    finally:
        close_connection(conn)


@app.route('/api/celebration-animation', methods=['GET'])
@jwt_required()
def get_celebration_animation():
    """Public (jwt) endpoint — used by the CRM on load."""
    conn = get_db_connection()
    if not conn:
        return jsonify({'error': 'Database connection failed'}), 500
    try:
        cursor = conn.cursor()
        cursor.execute(
            "SELECT setting_key, setting_value FROM app_settings WHERE setting_key IN (%s, %s)",
            ('celebration_animation_won', 'celebration_animation_lost')
        )
        rows = {r['setting_key']: r['setting_value'] for r in rows_to_list(cursor)}
        return jsonify({
            'won':  rows.get('celebration_animation_won',  'confetti'),
            'lost': rows.get('celebration_animation_lost', 'confetti'),
        })
    finally:
        close_connection(conn)


if __name__ == '__main__':
    debug_mode = os.getenv('FLASK_DEBUG', 'False').lower() == 'true'
    port = int(os.getenv('FLASK_PORT', '5000'))
    app.run(debug=debug_mode, port=port)
